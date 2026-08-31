from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import re
import statistics
from datetime import date, datetime, timedelta
from pathlib import Path
from uuid import uuid4
from zoneinfo import ZoneInfo

from app.db.database import get_connection


MAX_IMPORT_BYTES = 10 * 1024 * 1024
MAX_IMPORT_ROWS = 10_000
MAX_IMPORT_COLUMNS = 50
PREVIEW_TTL_HOURS = 24
DAILY_SOURCE_KIND = "account_daily_file"
DOUYIN_ITEM_EXPORT_SOURCE_KIND = "douyin_item_export"
BEIJING_TIMEZONE = ZoneInfo("Asia/Shanghai")

DAILY_HEADERS = (
    "日期",
    "投稿量",
    "总播放量",
    "总点赞量",
    "总分享量",
    "总评论量",
    "5秒完播率",
    "2秒跳出率",
    "封面点击率",
    "平均播放时长",
)

DOUYIN_ITEM_EXPORT_HEADERS = (
    "作品名称",
    "发布时间",
    "体裁",
    "审核状态",
    "播放量",
    "完播率",
    "5s完播率",
    "封面点击率",
    "2s跳出率",
    "平均播放时长",
    "点赞量",
    "分享量",
    "评论量",
    "收藏量",
    "主页访问量",
    "粉丝增量",
)

DOUYIN_ITEM_EXPORT_COUNT_FIELDS = {
    "播放量": "play_count",
    "点赞量": "like_count",
    "分享量": "share_count",
    "评论量": "comment_count",
    "收藏量": "collect_count",
    "主页访问量": "home_visit_count",
    "粉丝增量": "follower_gain_count",
}
DOUYIN_ITEM_EXPORT_RATE_FIELDS = {
    "完播率": "completion_rate",
    "5s完播率": "five_second_completion_rate",
    "封面点击率": "cover_click_rate",
    "2s跳出率": "two_second_bounce_rate",
}
DOUYIN_ITEM_EXPORT_FIELDS = (
    "aweme_id",
    "title",
    "published_at",
    "duration_seconds",
    "content_genre",
    "audit_status",
    "play_count",
    "completion_rate",
    "five_second_completion_rate",
    "cover_click_rate",
    "two_second_bounce_rate",
    "average_watch_seconds",
    "like_count",
    "share_count",
    "comment_count",
    "collect_count",
    "home_visit_count",
    "follower_gain_count",
)

COUNT_FIELDS = {
    "投稿量": "post_count",
    "总播放量": "play_count",
    "总点赞量": "like_count",
    "总分享量": "share_count",
    "总评论量": "comment_count",
}
RATE_FIELDS = {
    "5秒完播率": "five_second_completion_rate",
    "2秒跳出率": "two_second_bounce_rate",
    "封面点击率": "cover_click_rate",
}
REJECT_REASON_LABELS = {
    "not_funny": "不好笑",
    "missing_setup": "铺垫缺失",
    "fragmented": "片段不完整",
    "duplicate": "重复",
    "dragging": "节奏拖沓",
    "other": "其他",
    "worth_publishing": "保留",
}
MATCHED_STATUSES = {"matched_exact", "matched_unique", "confirmed_manual"}
MATCH_STATUS_LABELS = {
    "matched_exact": "已匹配·唯一证据",
    "matched_unique": "已匹配·唯一证据",
    "confirmed_manual": "已匹配·人工确认",
    "ambiguous": "待人工确认·多个候选",
    "unmatched": "未匹配·没有候选",
}


class ContentReviewError(ValueError):
    def __init__(self, message: str, *, status_code: int = 400):
        super().__init__(message)
        self.status_code = status_code


def _now() -> datetime:
    return datetime.now().astimezone()


def _now_iso() -> str:
    return _now().astimezone(BEIJING_TIMEZONE).isoformat(timespec="seconds")


def _resolve_douyin_account_id(account_id: str = "") -> str:
    normalized = str(account_id or "").strip()
    with get_connection() as connection:
        if normalized:
            row = connection.execute(
                "SELECT id FROM publish_accounts WHERE id = ? AND platform = 'douyin'",
                (normalized,),
            ).fetchone()
            if row is None:
                raise ContentReviewError("没有找到这个抖音账号，请先在发送中心添加账号")
            return normalized
        rows = connection.execute(
            "SELECT id FROM publish_accounts WHERE platform = 'douyin' ORDER BY created_at, id"
        ).fetchall()
    if len(rows) == 1:
        return str(rows[0]["id"])
    if not rows:
        raise ContentReviewError("还没有抖音账号，请先在发送中心添加账号")
    raise ContentReviewError("存在多个抖音账号，请先选择本次数据对应的账号")


def list_douyin_accounts() -> list[dict]:
    with get_connection() as connection:
        rows = connection.execute(
            """
            SELECT id, account_name, account_uid, login_status, login_checked_at
            FROM publish_accounts
            WHERE platform = 'douyin'
            ORDER BY created_at, id
            """
        ).fetchall()
    return [dict(row) for row in rows]


def _normalize_header(value) -> str:
    return re.sub(r"\s+", "", str(value or "").replace("\ufeff", "").strip())


def _trim_row(row: tuple | list) -> list:
    values = list(row)
    while values and (values[-1] is None or str(values[-1]).strip() == ""):
        values.pop()
    return values


def _load_xlsx_report(content: bytes) -> tuple[str, list[list]]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise ContentReviewError("Excel 文件无法读取，可能已损坏或不是有效的 .xlsx 文件") from exc

    matching_sheets: list[tuple[str, str, list[list]]] = []
    try:
        for worksheet in workbook.worksheets:
            if int(worksheet.max_column or 0) > MAX_IMPORT_COLUMNS:
                raise ContentReviewError(f"工作表“{worksheet.title}”超过 {MAX_IMPORT_COLUMNS} 列限制")
            rows: list[list] = []
            for raw_row in worksheet.iter_rows(values_only=True):
                row = _trim_row(raw_row)
                if not row or all(value is None or str(value).strip() == "" for value in row):
                    continue
                rows.append(row)
                if len(rows) > MAX_IMPORT_ROWS + 1:
                    raise ContentReviewError(f"工作表“{worksheet.title}”超过 {MAX_IMPORT_ROWS} 行数据限制")
            if not rows:
                continue
            normalized_headers = [_normalize_header(value) for value in rows[0]]
            normalized_set = set(normalized_headers)
            if (
                len(normalized_headers) == len(DOUYIN_ITEM_EXPORT_HEADERS)
                and len(normalized_set) == len(DOUYIN_ITEM_EXPORT_HEADERS)
                and normalized_set == set(DOUYIN_ITEM_EXPORT_HEADERS)
            ):
                matching_sheets.append((worksheet.title, DOUYIN_ITEM_EXPORT_SOURCE_KIND, rows))
            elif set(DAILY_HEADERS) <= normalized_set:
                matching_sheets.append((worksheet.title, DAILY_SOURCE_KIND, rows))
    finally:
        workbook.close()

    if not matching_sheets:
        raise ContentReviewError("没有找到可识别的抖音账号趋势表或官方 16 列作品列表表")
    if len(matching_sheets) > 1:
        names = "、".join(name for name, _, _ in matching_sheets)
        raise ContentReviewError(f"发现多个可导入工作表（{names}），请只保留一个数据工作表")
    _, source_kind, rows = matching_sheets[0]
    return source_kind, rows


def _load_csv_rows(content: bytes) -> list[list]:
    decoded = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            decoded = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise ContentReviewError("CSV 编码无法识别，请使用 UTF-8 或 GB18030")
    try:
        reader = csv.reader(io.StringIO(decoded, newline=""))
        rows = [_trim_row(row) for row in reader]
    except csv.Error as exc:
        raise ContentReviewError("CSV 文件格式损坏，无法读取") from exc
    rows = [row for row in rows if row and any(str(value).strip() for value in row)]
    if not rows:
        raise ContentReviewError("CSV 文件没有数据")
    if len(rows) - 1 > MAX_IMPORT_ROWS:
        raise ContentReviewError(f"CSV 超过 {MAX_IMPORT_ROWS} 行数据限制")
    if max(len(row) for row in rows) > MAX_IMPORT_COLUMNS:
        raise ContentReviewError(f"CSV 超过 {MAX_IMPORT_COLUMNS} 列限制")
    return rows


def _parse_date(value, *, row_number: int) -> str:
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    else:
        text = str(value or "").strip()
        parsed = None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
            try:
                parsed = datetime.strptime(text, fmt).date()
                break
            except ValueError:
                continue
        if parsed is None:
            raise ContentReviewError(f"第 {row_number} 行日期无法识别：{text or '空值'}")
    if parsed > _now().date():
        raise ContentReviewError(f"第 {row_number} 行日期晚于今天：{parsed.isoformat()}")
    return parsed.isoformat()


def _parse_count(value, *, header: str, row_number: int) -> int:
    text = ("" if value is None else str(value)).replace(",", "").strip()
    try:
        number = float(text)
    except (TypeError, ValueError) as exc:
        raise ContentReviewError(f"第 {row_number} 行“{header}”不是有效数字") from exc
    if not math.isfinite(number) or number < 0 or number > 1_000_000_000_000:
        raise ContentReviewError(f"第 {row_number} 行“{header}”超出合理范围")
    if not number.is_integer():
        raise ContentReviewError(f"第 {row_number} 行“{header}”必须是整数")
    return int(number)


def _parse_rate(value, *, header: str, row_number: int) -> float:
    text = ("" if value is None else str(value)).replace(",", "").strip()
    is_percent = text.endswith("%")
    if is_percent:
        text = text[:-1].strip()
    try:
        number = float(text)
    except (TypeError, ValueError) as exc:
        raise ContentReviewError(f"第 {row_number} 行“{header}”不是有效百分比") from exc
    if is_percent or number > 1:
        number /= 100
    if not math.isfinite(number) or not 0 <= number <= 1:
        raise ContentReviewError(f"第 {row_number} 行“{header}”必须在 0% 到 100% 之间")
    return round(number, 8)


def _parse_seconds(value, *, row_number: int) -> float:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        seconds = float(value)
    else:
        text = str(value or "").strip().lower().replace("秒", "").replace("s", "")
        try:
            seconds = float(text)
        except ValueError as exc:
            raise ContentReviewError(f"第 {row_number} 行“平均播放时长”不是有效秒数") from exc
    if not math.isfinite(seconds) or not 0 <= seconds <= 86_400:
        raise ContentReviewError(f"第 {row_number} 行“平均播放时长”超出合理范围")
    return round(seconds, 4)


def _is_empty_export_value(value) -> bool:
    return value is None or str(value).strip() in {"", "-"}


def _parse_optional_count(value, *, header: str, row_number: int) -> int | None:
    if _is_empty_export_value(value):
        return None
    return _parse_count(value, header=header, row_number=row_number)


def _parse_optional_rate(value, *, header: str, row_number: int) -> float | None:
    if _is_empty_export_value(value):
        return None
    return _parse_rate(value, header=header, row_number=row_number)


def _parse_optional_seconds(value, *, row_number: int) -> float | None:
    if _is_empty_export_value(value):
        return None
    return _parse_seconds(value, row_number=row_number)


def _parse_beijing_datetime(value, *, row_number: int) -> str:
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, datetime.min.time())
    else:
        text = str(value or "").strip()
        if not text or text == "-":
            raise ContentReviewError(f"第 {row_number} 行“发布时间”不能为空")
        parsed = None
        try:
            parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            pass
        if parsed is None:
            for fmt in (
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d %H:%M",
                "%Y/%m/%d %H:%M:%S",
                "%Y/%m/%d %H:%M",
                "%Y.%m.%d %H:%M:%S",
                "%Y.%m.%d %H:%M",
            ):
                try:
                    parsed = datetime.strptime(text, fmt)
                    break
                except ValueError:
                    continue
        if parsed is None:
            raise ContentReviewError(f"第 {row_number} 行“发布时间”无法识别：{text}")
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=BEIJING_TIMEZONE)
    else:
        parsed = parsed.astimezone(BEIJING_TIMEZONE)
    return parsed.isoformat(timespec="seconds")


def _optional_export_text(value) -> str | None:
    if _is_empty_export_value(value):
        return None
    return str(value).strip()


def _export_internal_key(title: str, published_at: str) -> str:
    identity = f"{normalize_title(title)}|{published_at}"
    return "export:" + hashlib.sha256(identity.encode("utf-8")).hexdigest()


def _normalize_douyin_item_export_rows(rows: list[list]) -> list[dict]:
    if len(rows) < 2:
        raise ContentReviewError("作品列表文件只有表头，没有可导入的作品")
    normalized_headers = [_normalize_header(value) for value in rows[0]]
    if (
        len(normalized_headers) != len(DOUYIN_ITEM_EXPORT_HEADERS)
        or len(set(normalized_headers)) != len(DOUYIN_ITEM_EXPORT_HEADERS)
        or set(normalized_headers) != set(DOUYIN_ITEM_EXPORT_HEADERS)
    ):
        raise ContentReviewError("作品列表必须严格包含官方 16 列表头，且不能增加、缺少或重复列")
    header_index = {
        header: normalized_headers.index(header)
        for header in DOUYIN_ITEM_EXPORT_HEADERS
    }
    result: list[dict] = []
    seen_ids: set[str] = set()
    for row_number, row in enumerate(rows[1:], start=2):
        if not row or all(_is_empty_export_value(value) for value in row):
            continue

        def value_for(header: str):
            index = header_index[header]
            return row[index] if index < len(row) else None

        title = str(value_for("作品名称") or "").strip()
        if not title or title == "-":
            raise ContentReviewError(f"第 {row_number} 行“作品名称”不能为空")
        published_at = _parse_beijing_datetime(value_for("发布时间"), row_number=row_number)
        aweme_id = _export_internal_key(title, published_at)
        if aweme_id in seen_ids:
            continue
        seen_ids.add(aweme_id)
        item = {
            "aweme_id": aweme_id,
            "title": title,
            "published_at": published_at,
            "duration_seconds": None,
            "content_genre": _optional_export_text(value_for("体裁")),
            "audit_status": _optional_export_text(value_for("审核状态")),
        }
        for header, field in DOUYIN_ITEM_EXPORT_COUNT_FIELDS.items():
            item[field] = _parse_optional_count(
                value_for(header),
                header=header,
                row_number=row_number,
            )
        for header, field in DOUYIN_ITEM_EXPORT_RATE_FIELDS.items():
            item[field] = _parse_optional_rate(
                value_for(header),
                header=header,
                row_number=row_number,
            )
        item["average_watch_seconds"] = _parse_optional_seconds(
            value_for("平均播放时长"),
            row_number=row_number,
        )
        result.append(item)
    if not result:
        raise ContentReviewError("作品列表中没有可导入的有效作品")
    result.sort(key=lambda item: (item["published_at"], item["aweme_id"]), reverse=True)
    return result


def parse_douyin_item_export(content: bytes) -> list[dict]:
    """解析抖音创作者中心官方作品列表；不保存原始文件或浏览器数据。"""
    if not content:
        raise ContentReviewError("下载的作品列表为空")
    if len(content) > MAX_IMPORT_BYTES:
        raise ContentReviewError("作品列表超过 10MB 限制")
    source_kind, rows = _load_xlsx_report(content)
    if source_kind != DOUYIN_ITEM_EXPORT_SOURCE_KIND:
        raise ContentReviewError("下载文件不是抖音官方 16 列作品列表")
    return _normalize_douyin_item_export_rows(rows)


def normalize_douyin_item_export_items(items: list[dict]) -> list[dict]:
    """重新校验 Worker 白名单数据，保证自动和人工导入使用完全相同的规范。"""
    if not isinstance(items, list) or not items:
        raise ContentReviewError("作品列表没有可导入数据")
    if len(items) > MAX_IMPORT_ROWS:
        raise ContentReviewError(f"作品列表超过 {MAX_IMPORT_ROWS} 行数据限制")
    normalized: list[dict] = []
    seen_ids: set[str] = set()
    for row_number, raw in enumerate(items, start=2):
        if not isinstance(raw, dict):
            raise ContentReviewError(f"第 {row_number} 行作品数据格式不正确")
        title = str(raw.get("title") or "").strip()
        if not title:
            raise ContentReviewError(f"第 {row_number} 行“作品名称”不能为空")
        published_at = _parse_beijing_datetime(raw.get("published_at"), row_number=row_number)
        aweme_id = _export_internal_key(title, published_at)
        if aweme_id in seen_ids:
            continue
        seen_ids.add(aweme_id)
        item = {
            "aweme_id": aweme_id,
            "title": title,
            "published_at": published_at,
            "duration_seconds": None,
            "content_genre": _optional_export_text(raw.get("content_genre")),
            "audit_status": _optional_export_text(raw.get("audit_status")),
        }
        for field in DOUYIN_ITEM_EXPORT_COUNT_FIELDS.values():
            item[field] = _parse_optional_count(
                raw.get(field),
                header=field,
                row_number=row_number,
            )
        for field in DOUYIN_ITEM_EXPORT_RATE_FIELDS.values():
            item[field] = _parse_optional_rate(
                raw.get(field),
                header=field,
                row_number=row_number,
            )
        item["average_watch_seconds"] = _parse_optional_seconds(
            raw.get("average_watch_seconds"),
            row_number=row_number,
        )
        normalized.append(item)
    if not normalized:
        raise ContentReviewError("作品列表中没有可导入的有效作品")
    normalized.sort(key=lambda item: (item["published_at"], item["aweme_id"]), reverse=True)
    return normalized


def _canonical_export_payload(items: list[dict]) -> tuple[list[dict], str, str]:
    normalized_items = normalize_douyin_item_export_items(items)
    payload_json = json.dumps(normalized_items, ensure_ascii=False, separators=(",", ":"))
    payload_hash = hashlib.sha256(payload_json.encode("utf-8")).hexdigest()
    return normalized_items, payload_json, payload_hash


def _normalize_daily_rows(rows: list[list]) -> list[dict]:
    if len(rows) < 2:
        raise ContentReviewError("文件只有表头，没有可导入的数据行")
    normalized_headers = [_normalize_header(value) for value in rows[0]]
    duplicates = sorted({header for header in normalized_headers if normalized_headers.count(header) > 1})
    if duplicates:
        raise ContentReviewError("表头存在重复列：" + "、".join(duplicates))
    missing = [header for header in DAILY_HEADERS if header not in normalized_headers]
    if missing:
        raise ContentReviewError("缺少必需表头：" + "、".join(missing))
    header_index = {header: normalized_headers.index(header) for header in DAILY_HEADERS}

    result: list[dict] = []
    seen_dates: set[str] = set()
    for row_number, row in enumerate(rows[1:], start=2):
        if not row or all(value is None or str(value).strip() == "" for value in row):
            continue

        def value_for(header: str):
            index = header_index[header]
            return row[index] if index < len(row) else None

        metric_date = _parse_date(value_for("日期"), row_number=row_number)
        if metric_date in seen_dates:
            raise ContentReviewError(f"日期重复：{metric_date}（第 {row_number} 行）")
        seen_dates.add(metric_date)
        item = {"metric_date": metric_date}
        for header, field in COUNT_FIELDS.items():
            item[field] = _parse_count(value_for(header), header=header, row_number=row_number)
        for header, field in RATE_FIELDS.items():
            item[field] = _parse_rate(value_for(header), header=header, row_number=row_number)
        item["average_watch_seconds"] = _parse_seconds(
            value_for("平均播放时长"),
            row_number=row_number,
        )
        result.append(item)
    if not result:
        raise ContentReviewError("没有可导入的有效数据行")
    result.sort(key=lambda item: item["metric_date"])
    return result


def preview_metric_import(*, account_id: str, filename: str, content: bytes) -> dict:
    resolved_account_id = _resolve_douyin_account_id(account_id)
    if not content:
        raise ContentReviewError("请选择要导入的数据文件")
    if len(content) > MAX_IMPORT_BYTES:
        raise ContentReviewError("文件超过 10MB 限制")
    safe_filename = Path(str(filename or "data")).name[:180]
    extension = Path(safe_filename).suffix.lower()
    if extension not in {".xlsx", ".csv"}:
        raise ContentReviewError("仅支持 .xlsx 和 .csv；不支持旧 .xls、宏文件或其他格式")
    if extension == ".xlsx":
        source_kind, rows = _load_xlsx_report(content)
    else:
        source_kind, rows = DAILY_SOURCE_KIND, _load_csv_rows(content)
    if source_kind == DOUYIN_ITEM_EXPORT_SOURCE_KIND:
        normalized_rows = _normalize_douyin_item_export_rows(rows)
        normalized_rows, normalized_json, source_sha256 = _canonical_export_payload(normalized_rows)
        published_dates = [str(item["published_at"])[:10] for item in normalized_rows]
        period_start = min(published_dates)
        period_end = max(published_dates)
        attribution = "item_level_pending_matching"
        message = "官方作品列表预览通过；确认后会写入作品快照并按唯一证据关联发布记录。"
        batch_prefix = "export"
        report_type = "douyin_item_export"
    else:
        normalized_rows = _normalize_daily_rows(rows)
        normalized_json = json.dumps(normalized_rows, ensure_ascii=False, separators=(",", ":"))
        source_sha256 = hashlib.sha256(content).hexdigest()
        period_start = normalized_rows[0]["metric_date"]
        period_end = normalized_rows[-1]["metric_date"]
        attribution = "unattributed_historical_baseline"
        message = "账号趋势表预览通过；确认后才会写入账号级历史基线。"
        batch_prefix = "metric"
        report_type = "account_daily"
    now = _now()
    now_iso = now.isoformat(timespec="seconds")
    expires_at = (now + timedelta(hours=PREVIEW_TTL_HOURS)).isoformat(timespec="seconds")

    with get_connection() as connection:
        connection.execute("BEGIN IMMEDIATE")
        existing = connection.execute(
            """
            SELECT id, status, created_at, committed_at, expires_at
            FROM content_metric_import_batches
            WHERE account_id = ? AND source_kind = ? AND source_sha256 = ?
            """,
            (resolved_account_id, source_kind, source_sha256),
        ).fetchone()
        if existing is not None and existing["status"] == "committed":
            connection.commit()
            return {
                "status": "already_imported",
                "already_imported": True,
                "message": "这个账号已经导入过相同的规范化数据，无需重复导入。",
                "batch_id": existing["id"],
                "account_id": resolved_account_id,
                "filename": safe_filename,
                "report_type": report_type,
                "row_count": len(normalized_rows),
                "period_start": period_start,
                "period_end": period_end,
            }
        if existing is not None:
            batch_id = str(existing["id"])
            connection.execute(
                """
                UPDATE content_metric_import_batches
                SET source_filename = ?, status = 'previewed', period_start = ?, period_end = ?,
                    normalized_payload_json = ?, row_count = ?, invalid_count = 0,
                    created_at = ?, committed_at = NULL, expires_at = ?
                WHERE id = ?
                """,
                (
                    safe_filename,
                    period_start,
                    period_end,
                    normalized_json,
                    len(normalized_rows),
                    now_iso,
                    expires_at,
                    batch_id,
                ),
            )
        else:
            batch_id = f"{batch_prefix}-{uuid4().hex[:16]}"
            connection.execute(
                """
                INSERT INTO content_metric_import_batches (
                    id, account_id, source_kind, source_filename, source_sha256, status,
                    period_start, period_end, normalized_payload_json, row_count,
                    invalid_count, created_at, expires_at
                ) VALUES (?, ?, ?, ?, ?, 'previewed', ?, ?, ?, ?, 0, ?, ?)
                """,
                (
                    batch_id,
                    resolved_account_id,
                    source_kind,
                    safe_filename,
                    source_sha256,
                    period_start,
                    period_end,
                    normalized_json,
                    len(normalized_rows),
                    now_iso,
                    expires_at,
                ),
            )
        connection.commit()
    return {
        "status": "previewed",
        "already_imported": False,
        "message": message,
        "batch_id": batch_id,
        "account_id": resolved_account_id,
        "filename": safe_filename,
        "report_type": report_type,
        "row_count": len(normalized_rows),
        "period_start": period_start,
        "period_end": period_end,
        "expires_at": expires_at,
        "sample_rows": normalized_rows[:5],
        "attribution": attribution,
    }


def commit_metric_import(batch_id: str) -> dict:
    now = _now()
    now_iso = now.isoformat(timespec="seconds")
    with get_connection() as connection:
        connection.execute("BEGIN IMMEDIATE")
        batch = connection.execute(
            "SELECT * FROM content_metric_import_batches WHERE id = ?",
            (batch_id,),
        ).fetchone()
        if batch is None:
            connection.rollback()
            raise ContentReviewError("导入预览不存在或已被清理", status_code=404)
        if batch["status"] == "committed":
            connection.commit()
            return {
                "status": "already_imported",
                "already_imported": True,
                "batch_id": batch_id,
                "message": "这批数据已经确认导入。",
            }
        if batch["status"] != "previewed":
            connection.rollback()
            raise ContentReviewError("这批预览当前不能确认导入", status_code=409)
        expires_at = datetime.fromisoformat(str(batch["expires_at"]))
        if expires_at < now:
            connection.execute(
                "UPDATE content_metric_import_batches SET status = 'expired' WHERE id = ?",
                (batch_id,),
            )
            connection.commit()
            raise ContentReviewError("预览已超过 24 小时，请重新选择文件预览", status_code=409)
        try:
            rows = json.loads(str(batch["normalized_payload_json"] or "[]"))
        except json.JSONDecodeError as exc:
            connection.rollback()
            raise ContentReviewError("预览数据损坏，请重新上传文件", status_code=409) from exc
        if not isinstance(rows, list) or not rows:
            connection.rollback()
            raise ContentReviewError("预览没有可导入数据，请重新上传文件", status_code=409)
        if batch["source_kind"] == DOUYIN_ITEM_EXPORT_SOURCE_KIND:
            normalized_items, _payload_json, payload_hash = _canonical_export_payload(rows)
            if payload_hash != str(batch["source_sha256"]):
                connection.rollback()
                raise ContentReviewError("作品预览规范化哈希校验失败，请重新上传文件", status_code=409)
            stats = _commit_export_batch_with_connection(
                connection,
                batch_id=str(batch["id"]),
                account_id=str(batch["account_id"]),
                items=normalized_items,
                captured_at=now_iso,
                committed_at=now_iso,
            )
            connection.commit()
            return {
                "status": "committed",
                "already_imported": False,
                "batch_id": batch_id,
                **stats,
                "message": (
                    f"已导入 {stats['row_count']} 条官方作品数据；"
                    f"唯一匹配 {stats['matched_count']} 条，歧义 {stats['ambiguous_count']} 条，"
                    f"未匹配 {stats['unmatched_count']} 条。"
                ),
                "attribution": "item_level_unique_only",
            }
        for row in rows:
            connection.execute(
                """
                INSERT INTO douyin_account_daily_metric_snapshots (
                    id, batch_id, account_id, metric_date, post_count, play_count,
                    like_count, share_count, comment_count, five_second_completion_rate,
                    two_second_bounce_rate, cover_click_rate, average_watch_seconds, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    f"daily-{uuid4().hex[:16]}",
                    batch_id,
                    batch["account_id"],
                    row["metric_date"],
                    row["post_count"],
                    row["play_count"],
                    row["like_count"],
                    row["share_count"],
                    row["comment_count"],
                    row["five_second_completion_rate"],
                    row["two_second_bounce_rate"],
                    row["cover_click_rate"],
                    row["average_watch_seconds"],
                    now_iso,
                ),
            )
        connection.execute(
            """
            UPDATE content_metric_import_batches
            SET status = 'committed', committed_at = ?, expires_at = NULL
            WHERE id = ?
            """,
            (now_iso, batch_id),
        )
        connection.commit()
    return {
        "status": "committed",
        "already_imported": False,
        "batch_id": batch_id,
        "row_count": len(rows),
        "message": f"已导入 {len(rows)} 天账号级数据，并标记为未归因历史基线。",
        "attribution": "unattributed_historical_baseline",
    }


def list_import_batches(account_id: str = "", limit: int = 20) -> list[dict]:
    resolved = _resolve_douyin_account_id(account_id)
    safe_limit = max(1, min(100, int(limit)))
    with get_connection() as connection:
        rows = connection.execute(
            """
            SELECT id, account_id, source_kind, source_filename, source_sha256, status,
                   period_start, period_end, row_count, matched_count, ambiguous_count,
                   invalid_count, created_at, committed_at, expires_at
            FROM content_metric_import_batches
            WHERE account_id = ?
            ORDER BY created_at DESC LIMIT ?
            """,
            (resolved, safe_limit),
        ).fetchall()
    return [dict(row) for row in rows]


def _aggregate_period(rows: list[dict]) -> dict:
    if not rows:
        return {
            "post_count": 0,
            "play_count": 0,
            "interaction_count": 0,
            "five_second_completion_rate": None,
            "two_second_bounce_rate": None,
            "cover_click_rate": None,
            "average_watch_seconds": None,
        }
    rate_fields = (
        "five_second_completion_rate",
        "two_second_bounce_rate",
        "cover_click_rate",
        "average_watch_seconds",
    )
    result = {
        "post_count": sum(int(row["post_count"] or 0) for row in rows),
        "play_count": sum(int(row["play_count"] or 0) for row in rows),
        "interaction_count": sum(
            int(row["like_count"] or 0)
            + int(row["share_count"] or 0)
            + int(row["comment_count"] or 0)
            for row in rows
        ),
    }
    for field in rate_fields:
        values = [float(row[field]) for row in rows if row[field] is not None]
        result[field] = round(sum(values) / len(values), 6) if values else None
    return result


def _comparison_delta(current, previous) -> float | None:
    if current is None or previous in (None, 0):
        return None
    return round((float(current) - float(previous)) / abs(float(previous)), 6)


def _parse_iso_datetime(value) -> datetime | None:
    normalized = str(value or "").strip()
    if not normalized:
        return None
    try:
        parsed = datetime.fromisoformat(normalized.replace("Z", "+00:00"))
        return parsed if parsed.tzinfo is not None else parsed.replace(tzinfo=BEIJING_TIMEZONE)
    except ValueError:
        return None


def _official_export_context(connection, account_id: str) -> dict:
    rows = connection.execute(
        """
        SELECT b.id, b.committed_at, b.row_count, b.matched_count,
               b.ambiguous_count, b.invalid_count,
               MAX(i.captured_at) AS captured_at
        FROM content_metric_import_batches b
        LEFT JOIN douyin_item_metric_snapshots i ON i.batch_id = b.id
        WHERE b.account_id = ? AND b.status = 'committed'
          AND b.source_kind = ?
        GROUP BY b.id
        ORDER BY b.committed_at DESC, b.created_at DESC
        """,
        (account_id, DOUYIN_ITEM_EXPORT_SOURCE_KIND),
    ).fetchall()
    week_keys = set()
    for row in rows:
        captured_at = _parse_iso_datetime(row["captured_at"] or row["committed_at"])
        if captured_at is None:
            continue
        localized = captured_at.astimezone(BEIJING_TIMEZONE)
        iso_year, iso_week, _ = localized.isocalendar()
        week_keys.add(f"{iso_year}-W{iso_week:02d}")
    latest = dict(rows[0]) if rows else {}
    row_count = int(latest.get("row_count") or 0)
    matched_count = int(latest.get("matched_count") or 0)
    ambiguous_count = int(latest.get("ambiguous_count") or 0)
    invalid_count = int(latest.get("invalid_count") or 0)
    return {
        "last_export_batch_id": latest.get("id"),
        "last_export_committed_at": latest.get("committed_at"),
        "last_export_captured_at": latest.get("captured_at"),
        "last_export_row_count": row_count,
        "last_export_matched_count": matched_count,
        "last_export_ambiguous_count": ambiguous_count,
        "last_export_unmatched_count": max(
            0,
            row_count - matched_count - ambiguous_count - invalid_count,
        ),
        "official_export_weeks": len(week_keys),
        "official_export_week_keys": sorted(week_keys),
    }


def _latest_match_summary(connection, account_id: str) -> dict:
    rows = connection.execute(
        """
        WITH latest_items AS (
            SELECT i.match_status,
                   ROW_NUMBER() OVER (
                       PARTITION BY CASE
                           WHEN i.publish_job_id IS NOT NULL THEN 'job:' || i.publish_job_id
                           ELSE 'work:' || i.aweme_id
                       END
                       ORDER BY i.captured_at DESC, i.created_at DESC, i.rowid DESC
                   ) AS item_rank
            FROM douyin_item_metric_snapshots i
            JOIN content_metric_import_batches b ON b.id = i.batch_id
            WHERE i.account_id = ? AND b.status = 'committed'
              AND b.source_kind = ?
        )
        SELECT match_status, COUNT(*) AS item_count
        FROM latest_items
        WHERE item_rank = 1
        GROUP BY match_status
        """,
        (account_id, DOUYIN_ITEM_EXPORT_SOURCE_KIND),
    ).fetchall()
    counts = {str(row["match_status"]): int(row["item_count"] or 0) for row in rows}
    return {
        "total": sum(counts.values()),
        "matched": sum(counts.get(status, 0) for status in MATCHED_STATUSES),
        "ambiguous": counts.get("ambiguous", 0),
        "unmatched": counts.get("unmatched", 0),
        "matched_exact": counts.get("matched_exact", 0),
        "matched_unique": counts.get("matched_unique", 0),
        "confirmed_manual": counts.get("confirmed_manual", 0),
    }


def get_content_review_summary(account_id: str = "", days: int = 28) -> dict:
    resolved = _resolve_douyin_account_id(account_id)
    safe_days = max(14, min(180, int(days)))
    with get_connection() as connection:
        rows = connection.execute(
            """
            WITH ranked AS (
                SELECT s.*, b.committed_at,
                       ROW_NUMBER() OVER (
                           PARTITION BY s.metric_date
                           ORDER BY b.committed_at DESC, s.created_at DESC, s.rowid DESC
                       ) AS metric_rank
                FROM douyin_account_daily_metric_snapshots s
                JOIN content_metric_import_batches b ON b.id = s.batch_id
                WHERE s.account_id = ? AND b.status = 'committed'
            )
            SELECT * FROM ranked WHERE metric_rank = 1 ORDER BY metric_date ASC
            """,
            (resolved,),
        ).fetchall()
        sync_row = connection.execute(
            """
            SELECT MAX(committed_at) AS last_sync_at
            FROM content_metric_import_batches
            WHERE account_id = ? AND status = 'committed'
            """,
            (resolved,),
        ).fetchone()
        export_context = _official_export_context(connection, resolved)
        match_summary = _latest_match_summary(connection, resolved)
    history = [dict(row) for row in rows]
    if not history:
        return {
            "account_id": resolved,
            "has_data": False,
            "message": "还没有确认导入的账号级数据。",
            "last_sync_at": None,
            "days_since_sync": None,
            "completed_cycles": export_context["official_export_weeks"],
            "current_period": _aggregate_period([]),
            "previous_period": _aggregate_period([]),
            "comparisons": {},
            "history": [],
            "match_summary": match_summary,
            **export_context,
        }

    latest_date = date.fromisoformat(history[-1]["metric_date"])
    current_start = latest_date - timedelta(days=6)
    previous_end = current_start - timedelta(days=1)
    previous_start = previous_end - timedelta(days=6)
    current_rows = [row for row in history if current_start <= date.fromisoformat(row["metric_date"]) <= latest_date]
    previous_rows = [
        row for row in history if previous_start <= date.fromisoformat(row["metric_date"]) <= previous_end
    ]
    current = _aggregate_period(current_rows)
    previous = _aggregate_period(previous_rows)
    comparisons = {
        key: _comparison_delta(current.get(key), previous.get(key))
        for key in current
    }
    last_sync_at = sync_row["last_sync_at"] if sync_row else None
    days_since = None
    if last_sync_at:
        days_since = max(0, (_now().date() - datetime.fromisoformat(str(last_sync_at)).date()).days)
    history_start = latest_date - timedelta(days=safe_days - 1)
    visible_history = [row for row in history if date.fromisoformat(row["metric_date"]) >= history_start]
    return {
        "account_id": resolved,
        "has_data": True,
        "latest_metric_date": latest_date.isoformat(),
        "last_sync_at": last_sync_at,
        "days_since_sync": days_since,
        "completed_cycles": export_context["official_export_weeks"],
        "current_period": current,
        "previous_period": previous,
        "comparisons": comparisons,
        "history": visible_history,
        "attribution": "unattributed_account_daily_baseline",
        "match_summary": match_summary,
        **export_context,
    }


def list_content_review_works(account_id: str = "", limit: int = 100) -> list[dict]:
    resolved = _resolve_douyin_account_id(account_id)
    safe_limit = max(1, min(200, int(limit)))
    with get_connection() as connection:
        rows = connection.execute(
            """
            WITH latest_items AS (
                SELECT i.*,
                       b.source_kind AS metric_source_kind,
                       b.source_filename AS metric_source_filename,
                       ROW_NUMBER() OVER (
                           PARTITION BY CASE
                               WHEN i.publish_job_id IS NOT NULL THEN 'job:' || i.publish_job_id
                               ELSE 'work:' || i.aweme_id
                           END
                           ORDER BY i.captured_at DESC, i.created_at DESC, i.rowid DESC
                       ) AS item_rank
                FROM douyin_item_metric_snapshots i
                JOIN content_metric_import_batches b ON b.id = i.batch_id
                WHERE i.account_id = ? AND b.status = 'committed'
            )
            SELECT i.*, pj.title AS publish_title, pj.published_at AS job_published_at,
                   pj.status AS publish_status, pj.output_clip_id, oc.clip_candidate_id,
                   c.title AS candidate_title, c.source_analysis_run_id,
                   ar.run_number AS analysis_run_number, ar.provider_label, ar.model,
                   ar.prompt_version_id, pv.version_number AS prompt_version_number,
                   pv.preset_name_snapshot AS prompt_name,
                   (
                       SELECT f.decision FROM clip_feedback f
                       WHERE f.clip_candidate_id = c.id
                       ORDER BY f.created_at DESC, f.rowid DESC LIMIT 1
                   ) AS review_decision,
                   (
                       SELECT f.reason_code FROM clip_feedback f
                       WHERE f.clip_candidate_id = c.id
                       ORDER BY f.created_at DESC, f.rowid DESC LIMIT 1
                   ) AS review_reason_code
            FROM latest_items i
            LEFT JOIN publish_jobs pj ON pj.id = i.publish_job_id
            LEFT JOIN output_clip oc ON oc.id = pj.output_clip_id
            LEFT JOIN clip_candidates c ON c.id = oc.clip_candidate_id
            LEFT JOIN ai_analysis_runs ar ON ar.id = c.source_analysis_run_id
            LEFT JOIN ai_prompt_versions pv ON pv.id = ar.prompt_version_id
            WHERE i.item_rank = 1
            ORDER BY COALESCE(i.published_at, i.captured_at) DESC
            LIMIT ?
            """,
            (resolved, safe_limit),
        ).fetchall()
    works = []
    for row in rows:
        work = dict(row)
        work["review_reason_label"] = REJECT_REASON_LABELS.get(
            str(work.get("review_reason_code") or ""),
            str(work.get("review_reason_code") or ""),
        )
        work["match_label"] = MATCH_STATUS_LABELS.get(
            str(work.get("match_status") or ""),
            "未匹配·没有候选",
        )
        work["attribution_complete"] = bool(
            work.get("publish_job_id")
            and work.get("clip_candidate_id")
            and work.get("source_analysis_run_id")
            and work.get("prompt_version_id")
        )
        works.append(work)
    return works


def _safe_ratio(numerator: float, denominator: float) -> float | None:
    if not denominator:
        return None
    return round(numerator / denominator, 6)


def get_prompt_comparison(account_id: str = "") -> dict:
    resolved = _resolve_douyin_account_id(account_id)
    with get_connection() as connection:
        candidate_rows = connection.execute(
            """
            SELECT pv.id AS prompt_version_id, pv.preset_id, pv.version_number,
                   pv.preset_name_snapshot, pv.created_at,
                   scoped.candidate_id, scoped.enabled, scoped.latest_decision
            FROM ai_prompt_versions pv
            JOIN (
                SELECT DISTINCT ar.prompt_version_id,
                       c.id AS candidate_id, c.enabled,
                       (
                           SELECT f.decision FROM clip_feedback f
                           WHERE f.clip_candidate_id = c.id
                           ORDER BY f.created_at DESC, f.rowid DESC LIMIT 1
                       ) AS latest_decision
                FROM ai_analysis_runs ar
                JOIN clip_candidates c
                  ON c.source_analysis_run_id = ar.id AND c.is_deleted = 0
                JOIN output_clip oc ON oc.clip_candidate_id = c.id
                JOIN publish_jobs pj ON pj.output_clip_id = oc.id
                WHERE pj.account_id = ? AND pj.platform = 'douyin'
                  AND ar.prompt_version_id IS NOT NULL
            ) scoped ON scoped.prompt_version_id = pv.id
            ORDER BY pv.created_at, pv.preset_id, pv.version_number
            """,
            (resolved,),
        ).fetchall()
        work_rows = connection.execute(
            """
            WITH latest_items AS (
                SELECT i.*,
                       ROW_NUMBER() OVER (
                           PARTITION BY CASE
                               WHEN i.publish_job_id IS NOT NULL THEN 'job:' || i.publish_job_id
                               ELSE 'work:' || i.aweme_id
                           END
                           ORDER BY i.captured_at DESC, i.created_at DESC, i.rowid DESC
                       ) AS item_rank
                FROM douyin_item_metric_snapshots i
                JOIN content_metric_import_batches b ON b.id = i.batch_id
                WHERE i.account_id = ? AND b.status = 'committed'
                  AND b.source_kind = ?
                  AND i.match_status IN ('matched_exact', 'matched_unique', 'confirmed_manual')
            )
            SELECT i.*, c.id AS candidate_id, ar.prompt_version_id
            FROM latest_items i
            JOIN publish_jobs pj ON pj.id = i.publish_job_id
            JOIN output_clip oc ON oc.id = pj.output_clip_id
            JOIN clip_candidates c ON c.id = oc.clip_candidate_id
            JOIN ai_analysis_runs ar ON ar.id = c.source_analysis_run_id
            WHERE i.item_rank = 1 AND ar.prompt_version_id IS NOT NULL
            """,
            (resolved, DOUYIN_ITEM_EXPORT_SOURCE_KIND),
        ).fetchall()
        export_context = _official_export_context(connection, resolved)

    groups: dict[str, dict] = {}
    for row in candidate_rows:
        version_id = str(row["prompt_version_id"])
        group = groups.setdefault(
            version_id,
            {
                "prompt_version_id": version_id,
                "preset_id": row["preset_id"],
                "version_number": int(row["version_number"]),
                "prompt_name": row["preset_name_snapshot"],
                "created_at": row["created_at"],
                "candidate_ids": set(),
                "kept_ids": set(),
                "published_candidate_ids": set(),
                "works": [],
            },
        )
        if row["candidate_id"]:
            candidate_id = str(row["candidate_id"])
            group["candidate_ids"].add(candidate_id)
            if row["latest_decision"] == "keep" or (
                row["latest_decision"] is None and int(row["enabled"] or 0) == 1
            ):
                group["kept_ids"].add(candidate_id)

    for row in work_rows:
        version_id = str(row["prompt_version_id"])
        group = groups.get(version_id)
        if group is None:
            continue
        group["works"].append(dict(row))
        group["published_candidate_ids"].add(str(row["candidate_id"]))

    versions = []
    for group in groups.values():
        candidate_count = len(group["candidate_ids"])
        works = group["works"]
        plays = [int(work["play_count"] or 0) for work in works]
        five_rates = [float(work["five_second_completion_rate"]) for work in works if work["five_second_completion_rate"] is not None]
        bounce_rates = [float(work["two_second_bounce_rate"]) for work in works if work["two_second_bounce_rate"] is not None]
        watch_ratios = [
            float(work["average_watch_seconds"]) / float(work["duration_seconds"])
            for work in works
            if work["average_watch_seconds"] is not None and float(work["duration_seconds"] or 0) > 0
        ]
        interactions = sum(
            int(work["like_count"] or 0)
            + int(work["comment_count"] or 0)
            + int(work["share_count"] or 0)
            for work in works
        )
        total_plays = sum(plays)
        versions.append(
            {
                "prompt_version_id": group["prompt_version_id"],
                "preset_id": group["preset_id"],
                "version_number": group["version_number"],
                "prompt_name": group["prompt_name"],
                "created_at": group["created_at"],
                "candidate_count": candidate_count,
                "accurate_published_count": len(works),
                "keep_rate": _safe_ratio(len(group["kept_ids"]), candidate_count),
                "publish_rate": _safe_ratio(len(group["published_candidate_ids"]), candidate_count),
                "median_play_count": statistics.median(plays) if plays else None,
                "five_second_completion_rate": statistics.fmean(five_rates) if five_rates else None,
                "two_second_bounce_rate": statistics.fmean(bounce_rates) if bounce_rates else None,
                "average_watch_ratio": statistics.fmean(watch_ratios) if watch_ratios else None,
                "interaction_rate": _safe_ratio(interactions, total_plays),
                "evaluable": len(works) >= 30,
                "required_accurate_published_count": 30,
                "remaining_accurate_published_count": max(0, 30 - len(works)),
            }
        )
    versions.sort(key=lambda item: (item["created_at"], item["preset_id"], item["version_number"]))
    comparisons = []
    for previous, current in zip(versions, versions[1:], strict=False):
        if previous["accurate_published_count"] < 20 or current["accurate_published_count"] < 20:
            continue
        comparisons.append(
            {
                "from_prompt_version_id": previous["prompt_version_id"],
                "to_prompt_version_id": current["prompt_version_id"],
                "note": "仅相关性，不代表因果；请结合选题、发布时间和样本结构判断。",
            }
        )
    completed_cycles = int(export_context["official_export_weeks"] or 0)
    return {
        "account_id": resolved,
        "completed_cycles": completed_cycles,
        "minimum_cycles": 3,
        "can_review_prompt": completed_cycles >= 3 and any(item["evaluable"] for item in versions),
        "versions": versions,
        "comparisons": comparisons,
        "message": (
            "已达到评估门槛，系统只提供建议，不会自动修改 Prompt。"
            if completed_cycles >= 3 and any(item["evaluable"] for item in versions)
            else "数据不足：需要 3 个不同官方导出周，且当前 Prompt 至少 30 条准确关联作品。"
        ),
        "causality_notice": "所有对比仅表示相关性，不代表因果。",
    }


DIAGNOSIS_CORE_METRICS = (
    "play_count",
    "five_second_completion_rate",
    "two_second_bounce_rate",
    "completion_rate",
    "watch_ratio",
)
EXPERIMENT_EDITABLE_JOB_STATUSES = {"DRAFT", "WAITING", "SCHEDULED"}
EXPERIMENT_DECISIONS = {"keep", "revert", "inconclusive", "cancel"}


def _duration_bucket(value) -> str:
    seconds = float(value or 0)
    if seconds <= 0:
        return "时长未知"
    if seconds <= 60:
        return "60秒内"
    if seconds <= 120:
        return "61-120秒"
    return "120秒以上"


def _publish_age_bucket(published_at, captured_at) -> str:
    published = _parse_iso_datetime(published_at)
    captured = _parse_iso_datetime(captured_at)
    if published is None or captured is None:
        return "发布年龄未知"
    days = max(0, (captured - published).days)
    if days <= 7:
        return "发布7天内"
    if days <= 30:
        return "发布8-30天"
    return "发布30天以上"


def _percentile(values: list[float], fraction: float) -> float | None:
    if not values:
        return None
    ordered = sorted(float(value) for value in values)
    if len(ordered) == 1:
        return ordered[0]
    position = (len(ordered) - 1) * fraction
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return ordered[lower]
    weight = position - lower
    return ordered[lower] * (1 - weight) + ordered[upper] * weight


def _cohort_benchmarks(rows: list[dict]) -> dict:
    metrics = {}
    for metric in DIAGNOSIS_CORE_METRICS:
        values = [float(row[metric]) for row in rows if row.get(metric) is not None]
        metrics[metric] = {
            "count": len(values),
            "p25": round(_percentile(values, 0.25), 6) if values else None,
            "median": round(statistics.median(values), 6) if values else None,
            "p75": round(_percentile(values, 0.75), 6) if values else None,
        }
    return metrics


def _latest_diagnosis_rows(connection, account_id: str) -> list[dict]:
    rows = connection.execute(
        """
        WITH latest_items AS (
            SELECT i.*, b.id AS metric_batch_id,
                   ROW_NUMBER() OVER (
                       PARTITION BY CASE
                           WHEN i.publish_job_id IS NOT NULL THEN 'job:' || i.publish_job_id
                           ELSE 'work:' || i.aweme_id
                       END
                       ORDER BY i.captured_at DESC, i.created_at DESC, i.rowid DESC
                   ) AS item_rank
            FROM douyin_item_metric_snapshots i
            JOIN content_metric_import_batches b ON b.id = i.batch_id
            WHERE i.account_id = ? AND b.status = 'committed'
              AND b.source_kind = ?
        )
        SELECT i.*, pj.title AS publish_title, oc.clip_candidate_id,
               c.title AS candidate_title,
               COALESCE(
                   NULLIF(i.duration_seconds, 0),
                   NULLIF(c.duration_seconds, 0),
                   NULLIF(oc.source_duration_ms, 0) / 1000.0
               ) AS effective_duration_seconds
        FROM latest_items i
        LEFT JOIN publish_jobs pj ON pj.id = i.publish_job_id
        LEFT JOIN output_clip oc ON oc.id = pj.output_clip_id
        LEFT JOIN clip_candidates c ON c.id = oc.clip_candidate_id
        WHERE i.item_rank = 1
        ORDER BY COALESCE(i.published_at, i.captured_at) DESC
        """,
        (account_id, DOUYIN_ITEM_EXPORT_SOURCE_KIND),
    ).fetchall()
    result = []
    for source in rows:
        row = dict(source)
        duration = float(row.get("effective_duration_seconds") or 0)
        average_watch = row.get("average_watch_seconds")
        row["watch_ratio"] = (
            round(float(average_watch) / duration, 6)
            if average_watch is not None and duration > 0
            else None
        )
        row["duration_bucket"] = _duration_bucket(duration)
        row["age_bucket"] = _publish_age_bucket(
            row.get("published_at"),
            row.get("captured_at"),
        )
        row["genre_bucket"] = str(row.get("content_genre") or "体裁未知")
        result.append(row)
    return result


def _select_comparable_cohort(work: dict, eligible: list[dict]) -> tuple[list[dict], dict]:
    selectors = (
        (
            "同体裁、同片长、同发布年龄",
            lambda item: item["genre_bucket"] == work["genre_bucket"]
            and item["duration_bucket"] == work["duration_bucket"]
            and item["age_bucket"] == work["age_bucket"],
        ),
        (
            "同体裁、同片长",
            lambda item: item["genre_bucket"] == work["genre_bucket"]
            and item["duration_bucket"] == work["duration_bucket"],
        ),
        (
            "同片长",
            lambda item: item["duration_bucket"] == work["duration_bucket"],
        ),
        ("全部准确匹配作品", lambda _item: True),
    )
    for label, selector in selectors:
        cohort = [item for item in eligible if selector(item)]
        if len(cohort) >= 8 or label == "全部准确匹配作品":
            return cohort, {
                "label": label,
                "genre": work["genre_bucket"],
                "duration": work["duration_bucket"],
                "publish_age": work["age_bucket"],
            }
    return eligible, {"label": "全部准确匹配作品"}


def _recommendation_definition(code: str) -> dict:
    definitions = {
        "weak_opening": {
            "title": "开头留存偏弱",
            "hypothesis": "前 1～2 秒的铺垫过长，用户尚未进入冲突、笑点或观点就离开。",
            "action_text": "下一批只改开头：删除开场铺垫，让前 1～2 秒直接进入冲突、笑点或核心观点。",
            "primary_metric": "two_second_bounce_rate",
            "primary_direction": "lower",
            "guardrail_metrics": ["five_second_completion_rate", "completion_rate"],
            "priority": 4,
        },
        "weak_pacing": {
            "title": "中段节奏偏弱",
            "hypothesis": "开头能够留住用户，但中段重复、停顿或解释拖慢了完播。",
            "action_text": "下一批只改中段：压缩重复、停顿和解释，并把关键结果提前。",
            "primary_metric": "completion_rate",
            "primary_direction": "higher",
            "guardrail_metrics": ["five_second_completion_rate", "two_second_bounce_rate"],
            "priority": 3,
        },
        "distribution_window": {
            "title": "留存尚可但播放偏低",
            "hypothesis": "片段本身的早期留存不差，低播放暂不能归咎于 Prompt。",
            "action_text": "保持选片、剪辑和文案不变，下一批只测试发布时间窗口。",
            "primary_metric": "play_count",
            "primary_direction": "higher",
            "guardrail_metrics": ["five_second_completion_rate", "two_second_bounce_rate"],
            "priority": 2,
        },
        "positive_reference": {
            "title": "可作为正样本",
            "hypothesis": "开头、持续观看和完播同时优于同类作品，可作为下批审片参考。",
            "action_text": "下一批只强化同类开场结构，其他选片条件保持不变。",
            "primary_metric": "five_second_completion_rate",
            "primary_direction": "higher",
            "guardrail_metrics": ["two_second_bounce_rate", "completion_rate"],
            "priority": 1,
        },
    }
    return definitions[code]


def _build_content_review_insights(account_id: str) -> dict:
    with get_connection() as connection:
        export_context = _official_export_context(connection, account_id)
        works = _latest_diagnosis_rows(connection, account_id)
    eligible = [
        work
        for work in works
        if work.get("publish_job_id")
        and str(work.get("match_status") or "") in MATCHED_STATUSES
        and all(work.get(metric) is not None for metric in DIAGNOSIS_CORE_METRICS)
    ]
    recommendations = []
    for work in eligible:
        cohort, cohort_key = _select_comparable_cohort(work, eligible)
        benchmarks = _cohort_benchmarks(cohort)
        five = float(work["five_second_completion_rate"])
        bounce = float(work["two_second_bounce_rate"])
        completion = float(work["completion_rate"])
        watch_ratio = float(work["watch_ratio"])
        plays = float(work["play_count"])
        code = ""
        score = 0.0
        if (
            bounce >= float(benchmarks["two_second_bounce_rate"]["p75"])
            and five <= float(benchmarks["five_second_completion_rate"]["p25"])
        ):
            code = "weak_opening"
            score = bounce - float(benchmarks["two_second_bounce_rate"]["p75"])
        elif five >= float(benchmarks["five_second_completion_rate"]["median"]) and (
            completion <= float(benchmarks["completion_rate"]["p25"])
            or watch_ratio <= float(benchmarks["watch_ratio"]["p25"])
        ):
            code = "weak_pacing"
            score = max(
                float(benchmarks["completion_rate"]["p25"]) - completion,
                float(benchmarks["watch_ratio"]["p25"]) - watch_ratio,
            )
        elif (
            plays <= float(benchmarks["play_count"]["p25"])
            and five >= float(benchmarks["five_second_completion_rate"]["median"])
            and bounce <= float(benchmarks["two_second_bounce_rate"]["median"])
        ):
            code = "distribution_window"
            score = (
                float(benchmarks["play_count"]["p25"]) - plays
            ) / max(float(benchmarks["play_count"]["p25"]), 1.0)
        elif (
            five >= float(benchmarks["five_second_completion_rate"]["p75"])
            and bounce <= float(benchmarks["two_second_bounce_rate"]["p25"])
            and completion >= float(benchmarks["completion_rate"]["p75"])
        ):
            code = "positive_reference"
            score = five - float(benchmarks["five_second_completion_rate"]["p75"])
        if not code:
            continue
        definition = _recommendation_definition(code)
        identity = "|".join(
            (
                account_id,
                str(export_context.get("last_export_batch_id") or ""),
                str(work["publish_job_id"]),
                code,
            )
        )
        recommendations.append(
            {
                "recommendation_id": hashlib.sha256(identity.encode("utf-8")).hexdigest()[:20],
                "diagnosis_code": code,
                "title": definition["title"],
                "hypothesis": definition["hypothesis"],
                "action_text": definition["action_text"],
                "primary_metric": definition["primary_metric"],
                "primary_direction": definition["primary_direction"],
                "guardrail_metrics": definition["guardrail_metrics"],
                "source_work": {
                    "publish_job_id": work["publish_job_id"],
                    "title": work.get("title") or work.get("publish_title") or "未命名作品",
                    "published_at": work.get("published_at"),
                },
                "evidence": {metric: work.get(metric) for metric in DIAGNOSIS_CORE_METRICS},
                "baseline": {
                    "batch_id": export_context.get("last_export_batch_id"),
                    "work_count": len(cohort),
                    "cohort": cohort_key,
                    "metrics": benchmarks,
                },
                "comparison_interval": {
                    metric: {
                        "p25": values.get("p25"),
                        "median": values.get("median"),
                        "p75": values.get("p75"),
                    }
                    for metric, values in benchmarks.items()
                },
                "data_sufficiency": "sufficient",
                "priority_score": round(float(definition["priority"]) + score, 6),
            }
        )
    recommendations.sort(key=lambda item: item["priority_score"], reverse=True)
    selected = []
    per_code = {}
    for item in recommendations:
        code = item["diagnosis_code"]
        if per_code.get(code, 0) >= 3:
            continue
        selected.append(item)
        per_code[code] = per_code.get(code, 0) + 1
        if len(selected) >= 12:
            break
    unmatched_count = sum(
        1
        for work in works
        if not work.get("publish_job_id")
        or str(work.get("match_status") or "") not in MATCHED_STATUSES
    )
    missing_metric_count = sum(
        1
        for work in works
        if work.get("publish_job_id")
        and str(work.get("match_status") or "") in MATCHED_STATUSES
        and any(work.get(metric) is None for metric in DIAGNOSIS_CORE_METRICS)
    )
    return {
        "account_id": account_id,
        "generated_at": _now_iso(),
        "baseline_batch_id": export_context.get("last_export_batch_id"),
        "official_export_weeks": export_context["official_export_weeks"],
        "summary": {
            "total_works": len(works),
            "eligible_works": len(eligible),
            "insufficient_works": len(works) - len(eligible),
            "unmatched_works": unmatched_count,
            "missing_metric_works": missing_metric_count,
            "recommendation_count": len(selected),
            "cover_metric_available": any(work.get("cover_click_rate") is not None for work in works),
            "note": "作品级封面点击率缺失时不会生成封面建议。",
        },
        "recommendations": selected,
    }


def _experiment_metric_summary(rows: list[dict]) -> dict:
    normalized = []
    for source in rows:
        row = dict(source)
        duration = float(row.get("effective_duration_seconds") or 0)
        row["watch_ratio"] = (
            float(row["average_watch_seconds"]) / duration
            if row.get("average_watch_seconds") is not None and duration > 0
            else None
        )
        normalized.append(row)
    return {
        metric: (
            round(statistics.median(values), 6)
            if (values := [float(row[metric]) for row in normalized if row.get(metric) is not None])
            else None
        )
        for metric in DIAGNOSIS_CORE_METRICS
    }


def _experiment_progress(connection, experiment: dict) -> dict:
    assigned_count = int(
        connection.execute(
            "SELECT COUNT(*) FROM content_improvement_experiment_items WHERE experiment_id = ?",
            (experiment["id"],),
        ).fetchone()[0]
    )
    rows = connection.execute(
        """
        WITH latest_items AS (
            SELECT i.*,
                   COALESCE(
                       NULLIF(i.duration_seconds, 0),
                       NULLIF(c.duration_seconds, 0),
                       NULLIF(oc.source_duration_ms, 0) / 1000.0
                   ) AS effective_duration_seconds,
                   ROW_NUMBER() OVER (
                       PARTITION BY i.publish_job_id
                       ORDER BY i.captured_at DESC, i.created_at DESC, i.rowid DESC
                   ) AS item_rank
            FROM content_improvement_experiment_items ei
            JOIN douyin_item_metric_snapshots i ON i.publish_job_id = ei.publish_job_id
            LEFT JOIN publish_jobs pj ON pj.id = i.publish_job_id
            LEFT JOIN output_clip oc ON oc.id = pj.output_clip_id
            LEFT JOIN clip_candidates c ON c.id = oc.clip_candidate_id
            JOIN content_metric_import_batches b ON b.id = i.batch_id
            WHERE ei.experiment_id = ? AND b.status = 'committed'
              AND b.source_kind = ?
              AND i.match_status IN ('matched_exact', 'matched_unique', 'confirmed_manual')
        )
        SELECT * FROM latest_items WHERE item_rank = 1
        """,
        (experiment["id"], DOUYIN_ITEM_EXPORT_SOURCE_KIND),
    ).fetchall()
    week_rows = connection.execute(
        """
        SELECT DISTINCT i.captured_at
        FROM content_improvement_experiment_items ei
        JOIN douyin_item_metric_snapshots i ON i.publish_job_id = ei.publish_job_id
        JOIN content_metric_import_batches b ON b.id = i.batch_id
        WHERE ei.experiment_id = ? AND b.status = 'committed'
          AND b.source_kind = ?
          AND i.match_status IN ('matched_exact', 'matched_unique', 'confirmed_manual')
        """,
        (experiment["id"], DOUYIN_ITEM_EXPORT_SOURCE_KIND),
    ).fetchall()
    weeks = set()
    for week_row in week_rows:
        captured = _parse_iso_datetime(week_row["captured_at"])
        if captured is None:
            continue
        year, week, _ = captured.astimezone(BEIJING_TIMEZONE).isocalendar()
        weeks.add(f"{year}-W{week:02d}")
    baseline = json.loads(str(experiment.get("baseline_json") or "{}"))
    baseline_count = int(baseline.get("work_count") or 0)
    treatment_count = len(rows)
    target = int(experiment.get("target_sample_size") or 20)
    minimum_baseline = int(experiment.get("minimum_baseline_size") or 20)
    minimum_weeks = int(experiment.get("minimum_weeks") or 3)
    decision_ready = (
        treatment_count >= target
        and baseline_count >= minimum_baseline
        and len(weeks) >= minimum_weeks
    )
    stage = "decision_ready" if decision_ready else ("early" if treatment_count >= 10 else "collecting")
    treatment_metrics = _experiment_metric_summary([dict(row) for row in rows])
    primary_metric = str(experiment.get("primary_metric") or "")
    baseline_primary = (
        baseline.get("metrics", {}).get(primary_metric, {}).get("median")
    )
    treatment_primary = treatment_metrics.get(primary_metric)
    primary_delta = (
        round(float(treatment_primary) - float(baseline_primary), 6)
        if treatment_primary is not None and baseline_primary is not None
        else None
    )
    return {
        "assigned_count": assigned_count,
        "treatment_count": treatment_count,
        "baseline_count": baseline_count,
        "official_export_weeks": len(weeks),
        "official_export_week_keys": sorted(weeks),
        "target_sample_size": target,
        "minimum_baseline_size": minimum_baseline,
        "minimum_weeks": minimum_weeks,
        "stage": stage,
        "trend_visible": treatment_count >= 10,
        "decision_ready": decision_ready,
        "treatment_metrics": treatment_metrics,
        "baseline_primary": baseline_primary,
        "treatment_primary": treatment_primary,
        "primary_delta": primary_delta,
    }


def list_content_experiments(account_id: str = "", *, include_closed: bool = True) -> list[dict]:
    resolved = _resolve_douyin_account_id(account_id)
    status_filter = "" if include_closed else " AND status = 'active'"
    with get_connection() as connection:
        rows = connection.execute(
            f"""
            SELECT * FROM content_improvement_experiments
            WHERE account_id = ?{status_filter}
            ORDER BY created_at DESC
            """,
            (resolved,),
        ).fetchall()
        experiments = []
        for source in rows:
            experiment = dict(source)
            experiment["guardrail_metrics"] = json.loads(
                str(experiment.pop("guardrail_metrics_json") or "[]")
            )
            experiment["baseline"] = json.loads(
                str(experiment.get("baseline_json") or "{}")
            )
            experiment["progress"] = _experiment_progress(connection, experiment)
            experiments.append(experiment)
    return experiments


def list_active_content_experiments_for_publish() -> list[dict]:
    with get_connection() as connection:
        rows = connection.execute(
            """
            SELECT id, account_id, title, action_text, primary_metric, created_at
            FROM content_improvement_experiments
            WHERE status = 'active'
            ORDER BY created_at DESC
            """
        ).fetchall()
    return [dict(row) for row in rows]


def get_content_review_insights(account_id: str = "") -> dict:
    resolved = _resolve_douyin_account_id(account_id)
    result = _build_content_review_insights(resolved)
    result["experiments"] = list_content_experiments(resolved)
    return result


def create_content_experiment(account_id: str, recommendation_id: str) -> dict:
    resolved = _resolve_douyin_account_id(account_id)
    insights = _build_content_review_insights(resolved)
    recommendation = next(
        (
            item
            for item in insights["recommendations"]
            if item["recommendation_id"] == str(recommendation_id or "").strip()
        ),
        None,
    )
    if recommendation is None:
        raise ContentReviewError("这条建议已经更新，请刷新页面后重新选择", status_code=409)
    baseline = recommendation["baseline"]
    if not baseline.get("batch_id"):
        raise ContentReviewError("还没有可冻结的官方作品导出基线", status_code=409)
    experiment_id = f"experiment-{uuid4().hex[:16]}"
    now = _now_iso()
    with get_connection() as connection:
        connection.execute("BEGIN IMMEDIATE")
        existing = connection.execute(
            """
            SELECT id FROM content_improvement_experiments
            WHERE account_id = ? AND recommendation_id = ?
            """,
            (resolved, recommendation["recommendation_id"]),
        ).fetchone()
        if existing is not None:
            connection.rollback()
            raise ContentReviewError("这条建议已经记录过实验，不能重复挑选同一批证据", status_code=409)
        connection.execute(
            """
            INSERT INTO content_improvement_experiments (
                id, account_id, recommendation_id, diagnosis_code, title,
                hypothesis, action_text, primary_metric, primary_direction,
                guardrail_metrics_json, baseline_batch_id, baseline_json,
                target_sample_size, minimum_baseline_size, minimum_weeks,
                status, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 20, 20, 3, 'active', ?, ?)
            """,
            (
                experiment_id,
                resolved,
                recommendation["recommendation_id"],
                recommendation["diagnosis_code"],
                recommendation["title"],
                recommendation["hypothesis"],
                recommendation["action_text"],
                recommendation["primary_metric"],
                recommendation["primary_direction"],
                json.dumps(recommendation["guardrail_metrics"], ensure_ascii=False),
                baseline["batch_id"],
                json.dumps(baseline, ensure_ascii=False),
                now,
                now,
            ),
        )
        connection.commit()
    return {
        "status": "created",
        "message": "实验已建立。请在发送中心投稿前标记实际采用该规则的作品。",
        "experiment_id": experiment_id,
        "experiment": next(
            item for item in list_content_experiments(resolved) if item["id"] == experiment_id
        ),
    }


def update_content_experiment(experiment_id: str, decision: str) -> dict:
    normalized = str(decision or "").strip().lower()
    if normalized not in EXPERIMENT_DECISIONS:
        raise ContentReviewError("实验结论只能是保留、回退、结论不足或取消")
    now = _now_iso()
    with get_connection() as connection:
        connection.execute("BEGIN IMMEDIATE")
        row = connection.execute(
            "SELECT * FROM content_improvement_experiments WHERE id = ?",
            (experiment_id,),
        ).fetchone()
        if row is None:
            connection.rollback()
            raise ContentReviewError("内容实验不存在", status_code=404)
        experiment = dict(row)
        if experiment["status"] != "active":
            connection.rollback()
            raise ContentReviewError("这个实验已经结束", status_code=409)
        progress = _experiment_progress(connection, experiment)
        if normalized != "cancel" and not progress["decision_ready"]:
            connection.rollback()
            raise ContentReviewError("样本或官方导出周数还不足，暂时不能记录最终结论", status_code=409)
        status = "cancelled" if normalized == "cancel" else "completed"
        stored_decision = None if normalized == "cancel" else normalized
        connection.execute(
            """
            UPDATE content_improvement_experiments
            SET status = ?, decision = ?, completed_at = ?, updated_at = ?
            WHERE id = ?
            """,
            (status, stored_decision, now, now, experiment_id),
        )
        connection.commit()
    return {"status": status, "decision": stored_decision, "message": "实验结论已记录。"}


def _assert_experiment_job_editable(job: dict) -> None:
    status = str(job.get("status") or "").upper()
    if (
        status not in EXPERIMENT_EDITABLE_JOB_STATUSES
        or job.get("claimed_at")
        or job.get("started_at")
        or int(job.get("attempt_count") or 0) > 0
    ):
        raise ContentReviewError("投稿执行已经开始，实验归属已冻结", status_code=409)


def assign_publish_job_to_experiment(experiment_id: str, publish_job_id: str) -> dict:
    now = _now_iso()
    with get_connection() as connection:
        connection.execute("BEGIN IMMEDIATE")
        experiment_row = connection.execute(
            "SELECT * FROM content_improvement_experiments WHERE id = ?",
            (experiment_id,),
        ).fetchone()
        if experiment_row is None:
            connection.rollback()
            raise ContentReviewError("内容实验不存在", status_code=404)
        experiment = dict(experiment_row)
        if experiment["status"] != "active":
            connection.rollback()
            raise ContentReviewError("只能关联进行中的实验", status_code=409)
        job_row = connection.execute(
            "SELECT * FROM publish_jobs WHERE id = ?",
            (publish_job_id,),
        ).fetchone()
        if job_row is None:
            connection.rollback()
            raise ContentReviewError("发布内容不存在", status_code=404)
        job = dict(job_row)
        _assert_experiment_job_editable(job)
        if job.get("platform") != "douyin" or job.get("account_id") != experiment["account_id"]:
            connection.rollback()
            raise ContentReviewError("实验与发布内容的抖音账号不一致", status_code=409)
        existing = connection.execute(
            """
            SELECT experiment_id FROM content_improvement_experiment_items
            WHERE publish_job_id = ?
            """,
            (publish_job_id,),
        ).fetchone()
        if existing is not None:
            if existing["experiment_id"] == experiment_id:
                connection.commit()
                return {"status": "already_assigned", "experiment_id": experiment_id}
            connection.rollback()
            raise ContentReviewError("这条作品已经属于另一个实验", status_code=409)
        connection.execute(
            """
            INSERT INTO content_improvement_experiment_items (
                id, experiment_id, publish_job_id, assigned_at
            ) VALUES (?, ?, ?, ?)
            """,
            (f"experiment-item-{uuid4().hex[:16]}", experiment_id, publish_job_id, now),
        )
        connection.commit()
    return {"status": "assigned", "experiment_id": experiment_id, "publish_job_id": publish_job_id}


def set_publish_job_experiment(publish_job_id: str, experiment_id: str = "") -> dict:
    """原子设置实验归属，避免前端先删旧关联、再加新关联时留下中间状态。"""

    normalized_experiment_id = str(experiment_id or "").strip()
    now = _now_iso()
    with get_connection() as connection:
        connection.execute("BEGIN IMMEDIATE")
        job_row = connection.execute(
            "SELECT * FROM publish_jobs WHERE id = ?",
            (publish_job_id,),
        ).fetchone()
        if job_row is None:
            connection.rollback()
            raise ContentReviewError("发布内容不存在", status_code=404)
        job = dict(job_row)
        _assert_experiment_job_editable(job)
        existing = connection.execute(
            """
            SELECT id, experiment_id
            FROM content_improvement_experiment_items
            WHERE publish_job_id = ?
            """,
            (publish_job_id,),
        ).fetchone()
        if existing is not None:
            current = connection.execute(
                "SELECT status FROM content_improvement_experiments WHERE id = ?",
                (existing["experiment_id"],),
            ).fetchone()
            if current is not None and current["status"] != "active":
                connection.rollback()
                raise ContentReviewError("实验已经结束，作品归属已冻结", status_code=409)
        if not normalized_experiment_id:
            if existing is not None:
                connection.execute(
                    "DELETE FROM content_improvement_experiment_items WHERE id = ?",
                    (existing["id"],),
                )
            connection.commit()
            return {"status": "removed" if existing is not None else "not_assigned", "publish_job_id": publish_job_id}

        experiment_row = connection.execute(
            "SELECT * FROM content_improvement_experiments WHERE id = ?",
            (normalized_experiment_id,),
        ).fetchone()
        if experiment_row is None:
            connection.rollback()
            raise ContentReviewError("内容实验不存在", status_code=404)
        experiment = dict(experiment_row)
        if experiment["status"] != "active":
            connection.rollback()
            raise ContentReviewError("只能关联进行中的实验", status_code=409)
        if job.get("platform") != "douyin" or job.get("account_id") != experiment["account_id"]:
            connection.rollback()
            raise ContentReviewError("实验与发布内容的抖音账号不一致", status_code=409)
        if existing is not None and existing["experiment_id"] == normalized_experiment_id:
            connection.commit()
            return {"status": "already_assigned", "experiment_id": normalized_experiment_id}
        if existing is None:
            connection.execute(
                """
                INSERT INTO content_improvement_experiment_items (
                    id, experiment_id, publish_job_id, assigned_at
                ) VALUES (?, ?, ?, ?)
                """,
                (
                    f"experiment-item-{uuid4().hex[:16]}",
                    normalized_experiment_id,
                    publish_job_id,
                    now,
                ),
            )
        else:
            connection.execute(
                """
                UPDATE content_improvement_experiment_items
                SET experiment_id = ?, assigned_at = ?
                WHERE id = ?
                """,
                (normalized_experiment_id, now, existing["id"]),
            )
        connection.commit()
    return {
        "status": "assigned" if existing is None else "reassigned",
        "experiment_id": normalized_experiment_id,
        "publish_job_id": publish_job_id,
    }


def remove_publish_job_from_experiment(experiment_id: str, publish_job_id: str) -> dict:
    with get_connection() as connection:
        connection.execute("BEGIN IMMEDIATE")
        item = connection.execute(
            """
            SELECT ei.id, e.status AS experiment_status
            FROM content_improvement_experiment_items ei
            JOIN content_improvement_experiments e ON e.id = ei.experiment_id
            WHERE ei.experiment_id = ? AND ei.publish_job_id = ?
            """,
            (experiment_id, publish_job_id),
        ).fetchone()
        if item is None:
            connection.commit()
            return {"status": "not_assigned"}
        job_row = connection.execute(
            "SELECT * FROM publish_jobs WHERE id = ?",
            (publish_job_id,),
        ).fetchone()
        if job_row is None:
            connection.rollback()
            raise ContentReviewError("发布内容不存在", status_code=404)
        _assert_experiment_job_editable(dict(job_row))
        if item["experiment_status"] != "active":
            connection.rollback()
            raise ContentReviewError("实验已经结束，作品归属已冻结", status_code=409)
        connection.execute(
            "DELETE FROM content_improvement_experiment_items WHERE id = ?",
            (item["id"],),
        )
        connection.commit()
    return {"status": "removed", "publish_job_id": publish_job_id}


def set_item_match(snapshot_id: str, publish_job_id: str) -> dict:
    now = _now_iso()
    with get_connection() as connection:
        connection.execute("BEGIN IMMEDIATE")
        snapshot = connection.execute(
            "SELECT id, account_id FROM douyin_item_metric_snapshots WHERE id = ?",
            (snapshot_id,),
        ).fetchone()
        if snapshot is None:
            connection.rollback()
            raise ContentReviewError("作品指标记录不存在", status_code=404)
        job = connection.execute(
            """
            SELECT id FROM publish_jobs
            WHERE id = ? AND platform = 'douyin' AND account_id = ?
            """,
            (publish_job_id, snapshot["account_id"]),
        ).fetchone()
        if job is None:
            connection.rollback()
            raise ContentReviewError("发布记录不存在，或不属于同一个抖音账号", status_code=409)
        connection.execute(
            """
            UPDATE douyin_item_metric_snapshots
            SET publish_job_id = ?, match_status = 'confirmed_manual',
                match_method = 'manual_confirmation', created_at = ?
            WHERE id = ?
            """,
            (publish_job_id, now, snapshot_id),
        )
        connection.commit()
    return {"status": "ok", "snapshot_id": snapshot_id, "publish_job_id": publish_job_id}


def delete_item_match(snapshot_id: str) -> dict:
    with get_connection() as connection:
        cursor = connection.execute(
            """
            UPDATE douyin_item_metric_snapshots
            SET publish_job_id = NULL, match_status = 'unmatched', match_method = NULL
            WHERE id = ?
            """,
            (snapshot_id,),
        )
        connection.commit()
    if cursor.rowcount != 1:
        raise ContentReviewError("作品指标记录不存在", status_code=404)
    return {"status": "ok", "snapshot_id": snapshot_id, "message": "已解除错误关联。"}


def normalize_title(value: str) -> str:
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", str(value or "").lower())


def _matching_datetime(value) -> datetime | None:
    try:
        parsed = datetime.fromisoformat(str(value or "").replace("Z", "+00:00"))
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=BEIJING_TIMEZONE)
    return parsed.astimezone(BEIJING_TIMEZONE)


def match_douyin_item_with_connection(connection, *, account_id: str, item: dict) -> dict:
    aweme_id = str(item.get("aweme_id") or "").strip()
    if aweme_id and not aweme_id.startswith("export:"):
        exact_rows = connection.execute(
            """
            SELECT id FROM publish_jobs
            WHERE platform = 'douyin' AND account_id = ?
              AND (platform_item_id = ? OR remote_video_id = ?)
            """,
            (account_id, aweme_id, aweme_id),
        ).fetchall()
        if len(exact_rows) == 1:
            return {
                "status": "matched_exact",
                "method": "platform_item_id",
                "publish_job_id": exact_rows[0]["id"],
            }
        if len(exact_rows) > 1:
            return {"status": "ambiguous", "method": "platform_item_id", "publish_job_id": None}

    normalized_title = normalize_title(str(item.get("title") or ""))
    item_time = _matching_datetime(item.get("published_at"))
    if not normalized_title or item_time is None:
        return {"status": "unmatched", "method": None, "publish_job_id": None}
    candidates = connection.execute(
        """
        SELECT pj.id, pj.title, pj.description, pj.caption, pj.published_at,
               oc.source_duration_ms
        FROM publish_jobs pj
        LEFT JOIN output_clip oc ON oc.id = pj.output_clip_id
        WHERE pj.platform = 'douyin' AND pj.account_id = ? AND pj.published_at IS NOT NULL
        """,
        (account_id,),
    ).fetchall()
    exact_matches: list = []
    contains_matches: list = []
    for candidate in candidates:
        candidate_time = _matching_datetime(candidate["published_at"])
        if candidate_time is None:
            continue
        seconds_apart = abs((candidate_time - item_time).total_seconds())
        if seconds_apart > 600:
            continue
        item_duration = float(item.get("duration_seconds") or 0)
        candidate_duration = float(candidate["source_duration_ms"] or 0) / 1000
        if item_duration and candidate_duration and abs(item_duration - candidate_duration) > 3:
            continue
        candidate_texts = {
            normalize_title(candidate[field])
            for field in ("title", "description", "caption")
            if normalize_title(candidate[field])
        }
        if normalized_title in candidate_texts:
            exact_matches.append(candidate)
            continue
        if len(normalized_title) < 8:
            continue
        if any(
            len(candidate_text) >= 8
            and (
                normalized_title in candidate_text
                or candidate_text in normalized_title
            )
            for candidate_text in candidate_texts
        ):
            contains_matches.append(candidate)
    if len(exact_matches) == 1:
        return {
            "status": "matched_unique",
            "method": "title_time_exact",
            "publish_job_id": exact_matches[0]["id"],
        }
    if len(exact_matches) > 1:
        return {"status": "ambiguous", "method": "title_time_exact", "publish_job_id": None}
    if len(contains_matches) == 1:
        return {
            "status": "matched_unique",
            "method": "title_time_contains",
            "publish_job_id": contains_matches[0]["id"],
        }
    if len(contains_matches) > 1:
        return {"status": "ambiguous", "method": "title_time_contains", "publish_job_id": None}
    return {"status": "unmatched", "method": None, "publish_job_id": None}


def _commit_export_batch_with_connection(
    connection,
    *,
    batch_id: str,
    account_id: str,
    items: list[dict],
    captured_at: str,
    committed_at: str,
) -> dict:
    matched = ambiguous = 0
    for item in items:
        match = match_douyin_item_with_connection(connection, account_id=account_id, item=item)
        if match["status"] in MATCHED_STATUSES:
            matched += 1
        elif match["status"] == "ambiguous":
            ambiguous += 1
        connection.execute(
            """
            INSERT INTO douyin_item_metric_snapshots (
                id, batch_id, publish_job_id, account_id, aweme_id, title,
                published_at, duration_seconds, captured_at, play_count, like_count,
                comment_count, share_count, collect_count, completion_rate,
                five_second_completion_rate, two_second_bounce_rate, cover_click_rate,
                average_watch_seconds, home_visit_count, follower_gain_count,
                content_genre, audit_status, match_status, match_method, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                f"item-{uuid4().hex[:16]}",
                batch_id,
                match["publish_job_id"],
                account_id,
                item["aweme_id"],
                item["title"],
                item["published_at"],
                item["duration_seconds"],
                captured_at,
                item["play_count"],
                item["like_count"],
                item["comment_count"],
                item["share_count"],
                item["collect_count"],
                item["completion_rate"],
                item["five_second_completion_rate"],
                item["two_second_bounce_rate"],
                item["cover_click_rate"],
                item["average_watch_seconds"],
                item["home_visit_count"],
                item["follower_gain_count"],
                item["content_genre"],
                item["audit_status"],
                match["status"],
                match["method"],
                committed_at,
            ),
        )
    connection.execute(
        """
        UPDATE content_metric_import_batches
        SET status = 'committed', matched_count = ?, ambiguous_count = ?, invalid_count = 0,
            committed_at = ?, expires_at = NULL
        WHERE id = ?
        """,
        (matched, ambiguous, committed_at, batch_id),
    )
    return {
        "row_count": len(items),
        "matched_count": matched,
        "ambiguous_count": ambiguous,
        "unmatched_count": len(items) - matched - ambiguous,
    }


def commit_douyin_item_export(
    *,
    account_id: str,
    items: list[dict],
    captured_at: str,
    source_filename: str,
) -> dict:
    resolved = _resolve_douyin_account_id(account_id)
    normalized_items, payload_json, payload_hash = _canonical_export_payload(items)
    safe_filename = Path(str(source_filename or "作品列表导出.xlsx")).name[:180]
    if Path(safe_filename).suffix.lower() != ".xlsx":
        safe_filename = "作品列表导出.xlsx"
    captured_iso = _parse_beijing_datetime(captured_at, row_number=0)
    now_iso = _now_iso()
    published_dates = [str(item["published_at"])[:10] for item in normalized_items]
    period_start = min(published_dates)
    period_end = max(published_dates)
    with get_connection() as connection:
        connection.execute("BEGIN IMMEDIATE")
        existing = connection.execute(
            """
            SELECT id, status, row_count, matched_count, ambiguous_count, source_filename
            FROM content_metric_import_batches
            WHERE account_id = ? AND source_kind = ? AND source_sha256 = ?
            """,
            (resolved, DOUYIN_ITEM_EXPORT_SOURCE_KIND, payload_hash),
        ).fetchone()
        if existing is not None and existing["status"] == "committed":
            row_count = int(existing["row_count"] or 0)
            matched_count = int(existing["matched_count"] or 0)
            ambiguous_count = int(existing["ambiguous_count"] or 0)
            connection.commit()
            return {
                "status": "already_imported",
                "already_imported": True,
                "batch_id": existing["id"],
                "source_filename": existing["source_filename"],
                "row_count": row_count,
                "matched_count": matched_count,
                "ambiguous_count": ambiguous_count,
                "unmatched_count": row_count - matched_count - ambiguous_count,
                "message": "官方作品数据没有变化，规范化内容已导入，无需重复保存。",
            }
        if existing is None:
            batch_id = f"export-{uuid4().hex[:16]}"
            connection.execute(
                """
                INSERT INTO content_metric_import_batches (
                    id, account_id, source_kind, source_filename, source_sha256, status,
                    period_start, period_end, normalized_payload_json, row_count,
                    created_at, committed_at
                ) VALUES (?, ?, ?, ?, ?, 'previewed', ?, ?, ?, ?, ?, NULL)
                """,
                (
                    batch_id,
                    resolved,
                    DOUYIN_ITEM_EXPORT_SOURCE_KIND,
                    safe_filename,
                    payload_hash,
                    period_start,
                    period_end,
                    payload_json,
                    len(normalized_items),
                    now_iso,
                ),
            )
        else:
            batch_id = str(existing["id"])
            connection.execute(
                """
                UPDATE content_metric_import_batches
                SET source_filename = ?, status = 'previewed', period_start = ?, period_end = ?,
                    normalized_payload_json = ?, row_count = ?, matched_count = 0,
                    ambiguous_count = 0, invalid_count = 0, created_at = ?,
                    committed_at = NULL, expires_at = NULL
                WHERE id = ?
                """,
                (
                    safe_filename,
                    period_start,
                    period_end,
                    payload_json,
                    len(normalized_items),
                    now_iso,
                    batch_id,
                ),
            )
        stats = _commit_export_batch_with_connection(
            connection,
            batch_id=batch_id,
            account_id=resolved,
            items=normalized_items,
            captured_at=captured_iso,
            committed_at=now_iso,
        )
        connection.commit()
    return {
        "status": "committed",
        "already_imported": False,
        "batch_id": batch_id,
        "source_filename": safe_filename,
        **stats,
        "message": (
            f"已同步 {stats['row_count']} 条官方作品数据；"
            f"唯一匹配 {stats['matched_count']} 条，歧义 {stats['ambiguous_count']} 条，"
            f"未匹配 {stats['unmatched_count']} 条。"
        ),
    }

from __future__ import annotations

import csv
from datetime import datetime
import hashlib
import io
import json
from uuid import uuid4

from fastapi.testclient import TestClient
from openpyxl import Workbook
import pytest

from app.db.database import get_connection, init_db
from app.main import app
from app.services import content_review_service
from app.services.publishers.base import PublishError
from app.services.publishers.worker_client import PublishWorkerClient


PREFIX = "test-content-review-"


@pytest.fixture(autouse=True)
def content_review_cleanup():
    init_db()
    _cleanup()
    yield
    _cleanup()


def _cleanup() -> None:
    with get_connection() as connection:
        account_rows = connection.execute(
            "SELECT id FROM publish_accounts WHERE id LIKE ?",
            (f"{PREFIX}%",),
        ).fetchall()
        account_ids = [row["id"] for row in account_rows]
        for account_id in account_ids:
            experiment_rows = connection.execute(
                "SELECT id FROM content_improvement_experiments WHERE account_id = ?",
                (account_id,),
            ).fetchall()
            for experiment in experiment_rows:
                connection.execute(
                    "DELETE FROM content_improvement_experiment_items WHERE experiment_id = ?",
                    (experiment["id"],),
                )
            connection.execute(
                "DELETE FROM content_improvement_experiments WHERE account_id = ?",
                (account_id,),
            )
            batch_rows = connection.execute(
                "SELECT id FROM content_metric_import_batches WHERE account_id = ?",
                (account_id,),
            ).fetchall()
            batch_ids = [row["id"] for row in batch_rows]
            for batch_id in batch_ids:
                connection.execute("DELETE FROM douyin_item_metric_snapshots WHERE batch_id = ?", (batch_id,))
                connection.execute(
                    "DELETE FROM douyin_account_daily_metric_snapshots WHERE batch_id = ?",
                    (batch_id,),
                )
            connection.execute("DELETE FROM content_metric_import_batches WHERE account_id = ?", (account_id,))
            job_rows = connection.execute(
                "SELECT id FROM publish_jobs WHERE account_id = ?",
                (account_id,),
            ).fetchall()
            for job in job_rows:
                connection.execute("DELETE FROM publish_job_events WHERE job_id = ?", (job["id"],))
            connection.execute("DELETE FROM publish_jobs WHERE account_id = ?", (account_id,))
        connection.execute("DELETE FROM output_clip WHERE task_id LIKE ?", (f"{PREFIX}%",))
        connection.execute("DELETE FROM clip_feedback WHERE task_id LIKE ?", (f"{PREFIX}%",))
        connection.execute("DELETE FROM clip_candidates WHERE task_id LIKE ?", (f"{PREFIX}%",))
        connection.execute("DELETE FROM ai_analysis_runs WHERE task_id LIKE ?", (f"{PREFIX}%",))
        connection.execute("DELETE FROM tasks WHERE id LIKE ?", (f"{PREFIX}%",))
        connection.execute("DELETE FROM ai_prompt_versions WHERE preset_id LIKE ?", (f"{PREFIX}%",))
        connection.execute("DELETE FROM ai_prompt_presets WHERE id LIKE ?", (f"{PREFIX}%",))
        connection.execute("DELETE FROM publish_accounts WHERE id LIKE ?", (f"{PREFIX}%",))
        connection.commit()


def _insert_account() -> str:
    account_id = f"{PREFIX}{uuid4().hex[:8]}"
    now = "2026-08-28T10:00:00+08:00"
    with get_connection() as connection:
        connection.execute(
            """
            INSERT INTO publish_accounts (
                id, platform, account_name, login_status, created_at, updated_at
            ) VALUES (?, 'douyin', '测试抖音账号', 'normal', ?, ?)
            """,
            (account_id, now, now),
        )
        connection.commit()
    return account_id


def _daily_rows() -> list[list]:
    return [
        ["2026-08-26", 2, 344, 7, 1, 0, "67.52%", "24.76%", "0.00%", "19.42s"],
        ["2026-08-27", 7, 1128, 45, 5, 2, 0.6159, 0.3014, 0.8571, 22.71],
    ]


def _xlsx_bytes(*, headers: list[str] | None = None, rows: list[list] | None = None) -> bytes:
    canonical_headers = list(content_review_service.DAILY_HEADERS)
    headers = headers or canonical_headers
    rows = rows or _daily_rows()
    order = [canonical_headers.index(header) for header in headers]
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row[index] for index in order])
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _csv_bytes(rows: list[list] | None = None) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(content_review_service.DAILY_HEADERS)
    writer.writerows(rows or _daily_rows())
    return output.getvalue().encode("utf-8-sig")


def _work_export_row(
    index: int,
    *,
    title: str | None = None,
    published_at: str | None = None,
) -> list:
    return [
        title or f"官方作品 {index:03d}",
        published_at or f"2026-08-{28 - (index % 20):02d} {20 - (index % 10):02d}:30",
        "视频",
        "审核通过",
        1000 + index,
        0.42,
        "58.5%",
        0.31,
        "-" if index == 1 else 0.12,
        "-" if index == 1 else 18.6,
        80 + index,
        12 + index,
        6 + index,
        9 + index,
        20 + index,
        3 + index,
    ]


def _work_export_xlsx(
    *,
    rows: list[list] | None = None,
    headers: list[str] | None = None,
    second_matching_sheet: bool = False,
) -> bytes:
    canonical_headers = list(content_review_service.DOUYIN_ITEM_EXPORT_HEADERS)
    headers = headers or canonical_headers
    rows = rows or [_work_export_row(1)]
    order = [canonical_headers.index(header) for header in headers if header in canonical_headers]
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row[index] for index in order])
    if second_matching_sheet:
        duplicate = workbook.create_sheet("Sheet2")
        duplicate.append(canonical_headers)
        duplicate.append(_work_export_row(99))
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_xlsx_preview_commit_summary_and_duplicate_are_safe():
    account_id = _insert_account()
    content = _xlsx_bytes(headers=list(reversed(content_review_service.DAILY_HEADERS)))

    preview = content_review_service.preview_metric_import(
        account_id=account_id,
        filename="抖音数据.xlsx",
        content=content,
    )
    committed = content_review_service.commit_metric_import(preview["batch_id"])
    duplicate = content_review_service.preview_metric_import(
        account_id=account_id,
        filename="重复.xlsx",
        content=content,
    )
    summary = content_review_service.get_content_review_summary(account_id, days=28)

    assert preview["row_count"] == 2
    assert preview["period_start"] == "2026-08-26"
    assert preview["sample_rows"][0]["five_second_completion_rate"] == 0.6752
    assert preview["sample_rows"][0]["comment_count"] == 0
    assert committed["attribution"] == "unattributed_historical_baseline"
    assert duplicate["already_imported"] is True
    assert summary["has_data"] is True
    assert summary["current_period"]["post_count"] == 9
    assert summary["current_period"]["play_count"] == 1472
    assert summary["attribution"] == "unattributed_account_daily_baseline"

    with get_connection() as connection:
        batch_columns = {
            row["name"] for row in connection.execute("PRAGMA table_info(content_metric_import_batches)")
        }
        stored_count = connection.execute(
            "SELECT COUNT(*) FROM douyin_account_daily_metric_snapshots WHERE account_id = ?",
            (account_id,),
        ).fetchone()[0]
    assert "file_content" not in batch_columns
    assert stored_count == 2


def test_csv_preview_supports_zero_percentage_and_seconds_suffix():
    account_id = _insert_account()
    preview = content_review_service.preview_metric_import(
        account_id=account_id,
        filename="account.csv",
        content=_csv_bytes(),
    )
    assert preview["sample_rows"][0]["cover_click_rate"] == 0
    assert preview["sample_rows"][0]["average_watch_seconds"] == 19.42


def test_official_work_export_parses_all_107_rows_and_full_16_columns():
    rows = [_work_export_row(index) for index in range(1, 108)]
    items = content_review_service.parse_douyin_item_export(
        _work_export_xlsx(rows=rows)
    )

    assert len(items) == 107
    assert all(item["aweme_id"].startswith("export:") for item in items)
    assert items == sorted(
        items,
        key=lambda item: (item["published_at"], item["aweme_id"]),
        reverse=True,
    )
    first_input = next(item for item in items if item["title"] == "官方作品 001")
    assert first_input["published_at"].endswith("+08:00")
    assert first_input["completion_rate"] == 0.42
    assert first_input["five_second_completion_rate"] == 0.585
    assert first_input["two_second_bounce_rate"] is None
    assert first_input["average_watch_seconds"] is None
    assert first_input["home_visit_count"] == 21
    assert first_input["follower_gain_count"] == 4
    assert first_input["content_genre"] == "视频"
    assert first_input["audit_status"] == "审核通过"


def test_official_work_export_manual_commit_and_auto_path_are_cross_idempotent():
    account_id = _insert_account()
    content = _work_export_xlsx(rows=[_work_export_row(1), _work_export_row(2)])
    preview = content_review_service.preview_metric_import(
        account_id=account_id,
        filename="作品列表导出.xlsx",
        content=content,
    )
    committed = content_review_service.commit_metric_import(preview["batch_id"])
    parsed_items = content_review_service.parse_douyin_item_export(content)
    duplicate = content_review_service.commit_douyin_item_export(
        account_id=account_id,
        items=parsed_items,
        captured_at="2026-08-29T10:00:00+08:00",
        source_filename="作品列表导出.xlsx",
    )

    assert preview["report_type"] == "douyin_item_export"
    assert committed["row_count"] == 2
    assert duplicate["status"] == "already_imported"
    assert duplicate["batch_id"] == preview["batch_id"]
    with get_connection() as connection:
        batch = connection.execute(
            "SELECT source_kind, source_sha256, row_count FROM content_metric_import_batches WHERE id = ?",
            (preview["batch_id"],),
        ).fetchone()
        snapshot = connection.execute(
            """
            SELECT completion_rate, five_second_completion_rate, home_visit_count,
                   follower_gain_count, content_genre, audit_status
            FROM douyin_item_metric_snapshots
            WHERE batch_id = ? AND title = '官方作品 001'
            """,
            (preview["batch_id"],),
        ).fetchone()
    assert batch["source_kind"] == "douyin_item_export"
    assert len(batch["source_sha256"]) == 64
    assert batch["row_count"] == 2
    assert dict(snapshot) == {
        "completion_rate": 0.42,
        "five_second_completion_rate": 0.585,
        "home_visit_count": 21,
        "follower_gain_count": 4,
        "content_genre": "视频",
        "audit_status": "审核通过",
    }


def test_official_work_export_auto_then_manual_is_cross_idempotent():
    account_id = _insert_account()
    content = _work_export_xlsx(rows=[_work_export_row(3)])
    parsed_items = content_review_service.parse_douyin_item_export(content)
    automatic = content_review_service.commit_douyin_item_export(
        account_id=account_id,
        items=parsed_items,
        captured_at="2026-08-29T10:00:00+08:00",
        source_filename="作品列表导出.xlsx",
    )
    manual = content_review_service.preview_metric_import(
        account_id=account_id,
        filename="人工下载的作品列表导出.xlsx",
        content=content,
    )
    assert automatic["status"] == "committed"
    assert manual["status"] == "already_imported"
    assert manual["batch_id"] == automatic["batch_id"]


@pytest.mark.parametrize(
    ("filename", "content", "message"),
    [
        ("bad.xlsx", b"not-a-workbook", "无法读取"),
        ("legacy.xls", b"legacy", "仅支持"),
        (
            "duplicate.csv",
            _csv_bytes(rows=[_daily_rows()[0], _daily_rows()[0]]),
            "日期重复",
        ),
    ],
)
def test_import_rejects_corrupt_legacy_and_duplicate_date(filename, content, message):
    account_id = _insert_account()
    with pytest.raises(content_review_service.ContentReviewError, match=message):
        content_review_service.preview_metric_import(
            account_id=account_id,
            filename=filename,
            content=content,
        )


def test_import_row_limit_and_wrong_sheet_are_rejected(monkeypatch):
    account_id = _insert_account()
    monkeypatch.setattr(content_review_service, "MAX_IMPORT_ROWS", 1)
    with pytest.raises(content_review_service.ContentReviewError, match="超过 1 行"):
        content_review_service.preview_metric_import(
            account_id=account_id,
            filename="too-many.xlsx",
            content=_xlsx_bytes(),
        )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["错误表头", "播放"])
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    with pytest.raises(content_review_service.ContentReviewError, match="可识别的抖音账号趋势表"):
        content_review_service.preview_metric_import(
            account_id=account_id,
            filename="wrong-sheet.xlsx",
            content=output.getvalue(),
        )


def test_official_work_export_rejects_wrong_headers_multiple_sheets_and_limits(monkeypatch):
    wrong_headers = list(content_review_service.DOUYIN_ITEM_EXPORT_HEADERS)
    wrong_headers[5] = "错误完播率"
    with pytest.raises(content_review_service.ContentReviewError, match="可识别"):
        content_review_service.parse_douyin_item_export(
            _work_export_xlsx(headers=wrong_headers)
        )
    with pytest.raises(content_review_service.ContentReviewError, match="多个可导入工作表"):
        content_review_service.parse_douyin_item_export(
            _work_export_xlsx(second_matching_sheet=True)
        )
    monkeypatch.setattr(content_review_service, "MAX_IMPORT_ROWS", 1)
    with pytest.raises(content_review_service.ContentReviewError, match="超过 1 行"):
        content_review_service.parse_douyin_item_export(
            _work_export_xlsx(rows=[_work_export_row(1), _work_export_row(2)])
        )


def test_import_rejects_file_over_10mb_before_parsing():
    account_id = _insert_account()
    with pytest.raises(content_review_service.ContentReviewError, match="10MB"):
        content_review_service.preview_metric_import(
            account_id=account_id,
            filename="too-large.xlsx",
            content=b"x" * (content_review_service.MAX_IMPORT_BYTES + 1),
        )


def _insert_publish_job(
    account_id: str,
    *,
    title: str,
    published_at: str,
    platform_item_id: str = "",
    duration_seconds: int = 60,
    description: str = "",
    caption: str = "",
    status: str = "PUBLISHED",
) -> str:
    suffix = uuid4().hex[:8]
    task_id = f"{PREFIX}task-{suffix}"
    output_id = f"{PREFIX}output-{suffix}"
    job_id = f"{PREFIX}job-{suffix}"
    now = "2026-08-28T10:00:00+08:00"
    with get_connection() as connection:
        connection.execute(
            "INSERT INTO tasks (id, task_name, task_dir_name, created_at, updated_at) VALUES (?, ?, ?, ?, ?)",
            (task_id, task_id, task_id, now, now),
        )
        connection.execute(
            """
            INSERT INTO output_clip (
                id, task_id, output_file_path, output_file_name, status,
                source_duration_ms, created_at, updated_at
            ) VALUES (?, ?, '', '', 'completed', ?, ?, ?)
            """,
            (output_id, task_id, duration_seconds * 1000, now, now),
        )
        connection.execute(
            """
            INSERT INTO publish_jobs (
                id, task_id, output_clip_id, account_id, platform, title, description,
                caption, status, platform_item_id, published_at, created_at, updated_at
            ) VALUES (?, ?, ?, ?, 'douyin', ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                job_id,
                task_id,
                output_id,
                account_id,
                title,
                description,
                caption,
                status,
                platform_item_id,
                published_at,
                now,
                now,
            ),
        )
        connection.commit()
    return job_id


def _diagnosis_item(index: int, *, title: str, published_at: str, **overrides) -> dict:
    item = {
        "aweme_id": f"export:test-{index}-{uuid4().hex[:6]}",
        "title": title,
        "published_at": published_at,
        "duration_seconds": 60,
        "content_genre": "视频",
        "audit_status": "审核通过",
        "play_count": 1000 + index * 100,
        "completion_rate": 0.45 + index * 0.005,
        "five_second_completion_rate": 0.60 + index * 0.005,
        "two_second_bounce_rate": 0.20 + index * 0.003,
        "average_watch_seconds": 30 + index,
        "cover_click_rate": None,
        "like_count": 20,
        "comment_count": 5,
        "share_count": 3,
    }
    item.update(overrides)
    return item


def _seed_diagnosis_baseline(account_id: str, *, captured_at: str = "2026-08-29T12:00:00+08:00") -> dict:
    items = []
    for index in range(9):
        published_at = f"2026-08-2{index}T10:00:00+08:00"
        title = f"诊断基线作品 {index}"
        _insert_publish_job(
            account_id,
            title=title,
            published_at=published_at,
            duration_seconds=60,
        )
        overrides = {}
        if index == 8:
            overrides = {
                "five_second_completion_rate": 0.30,
                "two_second_bounce_rate": 0.55,
                "completion_rate": 0.12,
                "average_watch_seconds": 12,
            }
        items.append(
            _diagnosis_item(
                index,
                title=title,
                published_at=published_at,
                **overrides,
            )
        )
    items.append(
        _diagnosis_item(
            99,
            title="没有候选的作品",
            published_at="2026-08-28T18:00:00+08:00",
        )
    )
    return content_review_service.commit_douyin_item_export(
        account_id=account_id,
        items=items,
        captured_at=captured_at,
        source_filename="作品列表导出.xlsx",
    )


def _attach_prompt_chain(job_id: str, label: str) -> str:
    suffix = uuid4().hex[:8]
    preset_id = f"{PREFIX}preset-{label}-{suffix}"
    version_id = f"{PREFIX}version-{label}-{suffix}"
    run_id = f"{PREFIX}run-{label}-{suffix}"
    candidate_id = f"{PREFIX}candidate-{label}-{suffix}"
    prompt_text = f"{label} prompt"
    prompt_hash = hashlib.sha256(prompt_text.encode("utf-8")).hexdigest()
    now = "2026-08-28T10:00:00+08:00"
    with get_connection() as connection:
        job = connection.execute(
            "SELECT task_id, output_clip_id FROM publish_jobs WHERE id = ?",
            (job_id,),
        ).fetchone()
        slot = int(connection.execute("SELECT COALESCE(MAX(slot), 1000) + 1 FROM ai_prompt_presets").fetchone()[0])
        connection.execute(
            """
            INSERT INTO ai_prompt_presets (
                id, slot, name, prompt_text, is_default, created_at, updated_at
            ) VALUES (?, ?, ?, ?, 0, ?, ?)
            """,
            (preset_id, slot, label, prompt_text, now, now),
        )
        connection.execute(
            """
            INSERT INTO ai_prompt_versions (
                id, preset_id, version_number, preset_name_snapshot,
                prompt_text, prompt_sha256, created_at
            ) VALUES (?, ?, 1, ?, ?, ?, ?)
            """,
            (version_id, preset_id, label, prompt_text, prompt_hash, now),
        )
        connection.execute(
            """
            INSERT INTO ai_analysis_runs (
                id, task_id, run_number, provider, provider_label, model,
                prompt_version_id, prompt_text_sha256, requested_clip_count,
                clip_count, analysis_payload_json, created_at
            ) VALUES (?, ?, 1, 'test', 'Test', 'test-model', ?, ?, 1, 1, ?, ?)
            """,
            (run_id, job["task_id"], version_id, prompt_hash, json.dumps({"clips": []}), now),
        )
        connection.execute(
            """
            INSERT INTO clip_candidates (
                id, task_id, clip_key, title, start_time, end_time,
                duration_seconds, source_analysis_run_id, created_at, updated_at
            ) VALUES (?, ?, 'clip_001', ?, '00:00:00', '00:01:00', 60, ?, ?, ?)
            """,
            (candidate_id, job["task_id"], label, run_id, now, now),
        )
        connection.execute(
            "UPDATE output_clip SET clip_candidate_id = ? WHERE id = ?",
            (candidate_id, job["output_clip_id"]),
        )
        connection.commit()
    return version_id


def test_item_matching_exact_unique_ambiguous_and_manual_confirmation():
    account_id = _insert_account()
    exact_job = _insert_publish_job(
        account_id,
        title="精确作品",
        published_at="2026-08-28T08:00:00+08:00",
        platform_item_id="aweme-exact",
    )
    unique_job = _insert_publish_job(
        account_id,
        title="唯一 标题！",
        published_at="2026-08-28T09:00:00+08:00",
    )
    ambiguous_a = _insert_publish_job(
        account_id,
        title="同名作品",
        published_at="2026-08-28T10:00:00+08:00",
    )
    _insert_publish_job(
        account_id,
        title="同名作品",
        published_at="2026-08-28T10:02:00+08:00",
    )
    with get_connection() as connection:
        exact = content_review_service.match_douyin_item_with_connection(
            connection,
            account_id=account_id,
            item={"aweme_id": "aweme-exact", "title": "x"},
        )
        unique = content_review_service.match_douyin_item_with_connection(
            connection,
            account_id=account_id,
            item={
                "aweme_id": "aweme-unique",
                "title": "唯一标题",
                "published_at": "2026-08-28T09:05:00+08:00",
                "duration_seconds": 60,
            },
        )
        ambiguous = content_review_service.match_douyin_item_with_connection(
            connection,
            account_id=account_id,
            item={
                "aweme_id": "aweme-ambiguous",
                "title": "同名作品",
                "published_at": "2026-08-28T10:01:00+08:00",
                "duration_seconds": 60,
            },
        )
    assert exact == {"status": "matched_exact", "method": "platform_item_id", "publish_job_id": exact_job}
    assert unique["status"] == "matched_unique"
    assert unique["publish_job_id"] == unique_job
    assert ambiguous["status"] == "ambiguous"

    sync = content_review_service.commit_douyin_item_export(
        account_id=account_id,
        captured_at="2026-08-28T11:00:00+08:00",
        source_filename="作品列表导出.xlsx",
        items=[
            {
                "title": "同名作品",
                "published_at": "2026-08-28T10:01:00+08:00",
                "play_count": 100,
            }
        ],
    )
    assert sync["ambiguous_count"] == 1
    works = content_review_service.list_content_review_works(account_id)
    assert works[0]["match_label"] == "待人工确认·多个候选"
    with get_connection() as connection:
        snapshot_id = connection.execute(
            "SELECT id FROM douyin_item_metric_snapshots WHERE title = '同名作品'"
        ).fetchone()[0]
    content_review_service.set_item_match(snapshot_id, ambiguous_a)
    works = content_review_service.list_content_review_works(account_id)
    assert works[0]["match_status"] == "confirmed_manual"
    assert works[0]["match_label"] == "已匹配·人工确认"
    assert works[0]["publish_job_id"] == ambiguous_a
    content_review_service.delete_item_match(snapshot_id)
    works = content_review_service.list_content_review_works(account_id)
    assert works[0]["match_status"] == "unmatched"


def test_export_matching_uses_unique_contains_but_protects_short_and_ambiguous_titles():
    account_id = _insert_account()
    contains_job = _insert_publish_job(
        account_id,
        title="发布标题",
        description="周杰伦谈童年往事完整版精彩片段",
        published_at="2026-08-28T09:00:00+08:00",
    )
    _insert_publish_job(
        account_id,
        title="短标题候选",
        description="康熙来了完整片段",
        published_at="2026-08-28T10:00:00+08:00",
    )
    _insert_publish_job(
        account_id,
        title="候选一",
        caption="小S蔡康永爆笑互怼名场面完整版一",
        published_at="2026-08-28T11:00:00+08:00",
    )
    _insert_publish_job(
        account_id,
        title="候选二",
        caption="小S蔡康永爆笑互怼名场面完整版二",
        published_at="2026-08-28T11:02:00+08:00",
    )
    with get_connection() as connection:
        contains = content_review_service.match_douyin_item_with_connection(
            connection,
            account_id=account_id,
            item={
                "aweme_id": "export:test",
                "title": "周杰伦谈童年往事完整版",
                "published_at": "2026-08-28T09:05:00+08:00",
            },
        )
        short = content_review_service.match_douyin_item_with_connection(
            connection,
            account_id=account_id,
            item={
                "aweme_id": "export:short",
                "title": "康熙",
                "published_at": "2026-08-28T10:03:00+08:00",
            },
        )
        ambiguous = content_review_service.match_douyin_item_with_connection(
            connection,
            account_id=account_id,
            item={
                "aweme_id": "export:ambiguous",
                "title": "小S蔡康永爆笑互怼名场面完整版",
                "published_at": "2026-08-28T11:01:00+08:00",
            },
        )
    assert contains == {
        "status": "matched_unique",
        "method": "title_time_contains",
        "publish_job_id": contains_job,
    }
    assert short["status"] == "unmatched"
    assert ambiguous["status"] == "ambiguous"


def test_latest_export_snapshot_is_grouped_by_publish_job_after_title_change():
    account_id = _insert_account()
    original_title = "周杰伦谈童年故事完整版"
    changed_title = "周杰伦谈童年故事完整版标题已修改"
    job_id = _insert_publish_job(
        account_id,
        title="平台发布标题",
        description=original_title,
        published_at="2026-08-28T09:00:00+08:00",
    )
    first = content_review_service.commit_douyin_item_export(
        account_id=account_id,
        items=[
            {
                "title": original_title,
                "published_at": "2026-08-28T09:00:00+08:00",
                "play_count": 100,
            }
        ],
        captured_at="2026-08-28T12:00:00+08:00",
        source_filename="作品列表导出.xlsx",
    )
    second = content_review_service.commit_douyin_item_export(
        account_id=account_id,
        items=[
            {
                "title": changed_title,
                "published_at": "2026-08-28T09:00:00+08:00",
                "play_count": 200,
            }
        ],
        captured_at="2026-08-29T12:00:00+08:00",
        source_filename="作品列表导出.xlsx",
    )
    works = content_review_service.list_content_review_works(account_id)
    assert first["matched_count"] == 1
    assert second["matched_count"] == 1
    assert len(works) == 1
    assert works[0]["publish_job_id"] == job_id
    assert works[0]["title"] == changed_title
    assert works[0]["play_count"] == 200


def test_content_review_page_and_import_api_flow():
    account_id = _insert_account()
    client = TestClient(app)
    page = client.get("/content-review")
    assert page.status_code == 200
    assert "内容复盘" in page.text
    assert "发送中心" in page.text
    assert page.text.index('href="/publish"') < page.text.index('href="/content-review"')
    assert page.text.index('href="/content-review"') < page.text.index('href="/system"')

    preview = client.post(
        "/api/content-review/imports/preview",
        data={"account_id": account_id},
        files={
            "file": (
                "daily.xlsx",
                _xlsx_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert preview.status_code == 200
    batch_id = preview.json()["batch_id"]
    committed = client.post(
        f"/api/content-review/imports/{batch_id}/commit",
        json={"confirm": True},
    )
    assert committed.status_code == 200
    summary = client.get(f"/api/content-review/summary?account_id={account_id}&days=28")
    assert summary.status_code == 200
    assert summary.json()["current_period"]["play_count"] == 1472


def test_export_sync_api_commits_sanitized_worker_items(monkeypatch):
    account_id = _insert_account()
    worker_items = content_review_service.parse_douyin_item_export(
        _work_export_xlsx(rows=[_work_export_row(1, title="同步测试")])
    )

    def worker_sync(_self, *, account_id):
        return {
            "captured_at": "2026-08-28T12:00:00+08:00",
            "source_filename": "作品列表导出.xlsx",
            "row_count": 1,
            "items": worker_items,
        }

    monkeypatch.setattr(PublishWorkerClient, "analytics_export_sync", worker_sync)
    response = TestClient(app).post(
        "/api/content-review/douyin/export-sync",
        json={"account_id": account_id},
    )
    assert response.status_code == 200
    assert response.json()["row_count"] == 1
    with get_connection() as connection:
        item = connection.execute(
            """
            SELECT aweme_id, title, play_count, like_count, completion_rate
            FROM douyin_item_metric_snapshots WHERE account_id = ?
            """,
            (account_id,),
        ).fetchone()
    assert dict(item) == {
        "aweme_id": worker_items[0]["aweme_id"],
        "title": "同步测试",
        "play_count": 1001,
        "like_count": 81,
        "completion_rate": 0.42,
    }


def test_export_sync_api_preserves_fixed_worker_error_code(monkeypatch):
    account_id = _insert_account()

    def rate_limited(_self, *, account_id):
        raise PublishError("rate limited", "RATE_LIMITED")

    monkeypatch.setattr(PublishWorkerClient, "analytics_export_sync", rate_limited)
    response = TestClient(app).post(
        "/api/content-review/douyin/export-sync",
        json={"account_id": account_id},
    )
    assert response.status_code == 429
    assert response.json()["detail"]["error_code"] == "RATE_LIMITED"


def test_summary_uses_latest_official_export_and_deduplicates_same_week(monkeypatch):
    account_id = _insert_account()
    job_id = _insert_publish_job(
        account_id,
        title="周口径测试作品",
        published_at="2026-08-28T10:00:00+08:00",
    )
    clock = {"value": datetime.fromisoformat("2026-08-29T00:48:37+08:00")}
    monkeypatch.setattr(content_review_service, "_now", lambda: clock["value"])
    first = content_review_service.commit_douyin_item_export(
        account_id=account_id,
        items=[
            _diagnosis_item(
                1,
                title="周口径测试作品",
                published_at="2026-08-28T10:00:00+08:00",
            )
        ],
        captured_at="2026-08-29T00:47:00+08:00",
        source_filename="作品列表导出.xlsx",
    )
    clock["value"] = datetime.fromisoformat("2026-08-30T08:15:00+08:00")
    second = content_review_service.commit_douyin_item_export(
        account_id=account_id,
        items=[
            _diagnosis_item(
                1,
                title="周口径测试作品",
                published_at="2026-08-28T10:00:00+08:00",
                play_count=2222,
            )
        ],
        captured_at="2026-08-30T08:14:00+08:00",
        source_filename="作品列表导出.xlsx",
    )
    clock["value"] = datetime.fromisoformat("2026-09-01T09:00:00+08:00")
    preview = content_review_service.preview_metric_import(
        account_id=account_id,
        filename="账号趋势.xlsx",
        content=_xlsx_bytes(),
    )
    content_review_service.commit_metric_import(preview["batch_id"])

    summary = content_review_service.get_content_review_summary(account_id)

    assert first["matched_count"] == second["matched_count"] == 1
    assert summary["last_export_batch_id"] == second["batch_id"]
    assert summary["last_export_committed_at"] == "2026-08-30T08:15:00+08:00"
    assert summary["last_export_row_count"] == 1
    assert summary["last_export_matched_count"] == 1
    assert summary["last_export_unmatched_count"] == 0
    assert summary["official_export_weeks"] == 1
    assert summary["completed_cycles"] == 1
    assert summary["last_sync_at"] == "2026-09-01T09:00:00+08:00"
    assert job_id


def test_diagnosis_uses_matched_metrics_and_skips_missing_cover_or_unmatched_works():
    account_id = _insert_account()
    export = _seed_diagnosis_baseline(account_id)

    insights = content_review_service.get_content_review_insights(account_id)
    works = content_review_service.list_content_review_works(account_id)

    assert export["matched_count"] == 9
    assert export["ambiguous_count"] == 0
    assert insights["summary"] == {
        "total_works": 10,
        "eligible_works": 9,
        "insufficient_works": 1,
        "unmatched_works": 1,
        "missing_metric_works": 0,
        "recommendation_count": len(insights["recommendations"]),
        "cover_metric_available": False,
        "note": "作品级封面点击率缺失时不会生成封面建议。",
    }
    weak_opening = next(
        item for item in insights["recommendations"] if item["diagnosis_code"] == "weak_opening"
    )
    assert weak_opening["data_sufficiency"] == "sufficient"
    assert weak_opening["source_work"]["title"] == "诊断基线作品 8"
    assert weak_opening["comparison_interval"]["two_second_bounce_rate"]["p75"] is not None
    assert all(item["source_work"]["title"] != "没有候选的作品" for item in insights["recommendations"])
    labels = {work["match_status"]: work["match_label"] for work in works}
    assert labels["matched_unique"] == "已匹配·唯一证据"
    assert labels["unmatched"] == "未匹配·没有候选"


def test_prompt_comparison_only_includes_versions_used_by_selected_account():
    first_account = _insert_account()
    second_account = _insert_account()
    first_job = _insert_publish_job(
        first_account,
        title="账号一作品",
        published_at="2026-08-28T10:00:00+08:00",
    )
    second_job = _insert_publish_job(
        second_account,
        title="账号二作品",
        published_at="2026-08-28T11:00:00+08:00",
    )
    first_version = _attach_prompt_chain(first_job, "账号一 Prompt")
    second_version = _attach_prompt_chain(second_job, "账号二 Prompt")

    first = content_review_service.get_prompt_comparison(first_account)
    second = content_review_service.get_prompt_comparison(second_account)

    assert [item["prompt_version_id"] for item in first["versions"]] == [first_version]
    assert [item["prompt_version_id"] for item in second["versions"]] == [second_version]
    assert first["completed_cycles"] == 0
    assert second["completed_cycles"] == 0


def test_experiment_assignment_is_unique_atomic_and_freezes_after_execution():
    account_id = _insert_account()
    _seed_diagnosis_baseline(account_id)
    insights = content_review_service.get_content_review_insights(account_id)
    recommendation = next(
        item for item in insights["recommendations"] if item["diagnosis_code"] == "weak_opening"
    )
    created = content_review_service.create_content_experiment(
        account_id,
        recommendation["recommendation_id"],
    )
    first_experiment_id = created["experiment_id"]
    assert created["experiment"]["progress"]["trend_visible"] is False
    second_experiment_id = f"{PREFIX}experiment-second-{uuid4().hex[:8]}"
    with get_connection() as connection:
        first = connection.execute(
            "SELECT * FROM content_improvement_experiments WHERE id = ?",
            (first_experiment_id,),
        ).fetchone()
        connection.execute(
            """
            INSERT INTO content_improvement_experiments (
                id, account_id, recommendation_id, diagnosis_code, title,
                hypothesis, action_text, primary_metric, primary_direction,
                guardrail_metrics_json, baseline_batch_id, baseline_json,
                status, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'active', ?, ?)
            """,
            (
                second_experiment_id,
                account_id,
                f"second-{uuid4().hex[:8]}",
                first["diagnosis_code"],
                "第二个实验",
                first["hypothesis"],
                first["action_text"],
                first["primary_metric"],
                first["primary_direction"],
                first["guardrail_metrics_json"],
                first["baseline_batch_id"],
                first["baseline_json"],
                first["created_at"],
                first["updated_at"],
            ),
        )
        connection.commit()
    job_id = _insert_publish_job(
        account_id,
        title="实验作品",
        published_at="2026-08-30T10:00:00+08:00",
        status="WAITING",
    )

    assigned = content_review_service.set_publish_job_experiment(job_id, first_experiment_id)
    reassigned = content_review_service.set_publish_job_experiment(job_id, second_experiment_id)
    assert assigned["status"] == "assigned"
    assert reassigned["status"] == "reassigned"
    with get_connection() as connection:
        count = connection.execute(
            "SELECT COUNT(*) FROM content_improvement_experiment_items WHERE publish_job_id = ?",
            (job_id,),
        ).fetchone()[0]
        owner = connection.execute(
            "SELECT experiment_id FROM content_improvement_experiment_items WHERE publish_job_id = ?",
            (job_id,),
        ).fetchone()[0]
        connection.execute(
            """
            UPDATE publish_jobs
            SET status = 'PUBLISHING', claimed_at = '2026-08-30T10:05:00+08:00',
                started_at = '2026-08-30T10:05:00+08:00', attempt_count = 1
            WHERE id = ?
            """,
            (job_id,),
        )
        connection.commit()
    assert count == 1
    assert owner == second_experiment_id
    with pytest.raises(content_review_service.ContentReviewError, match="已经开始"):
        content_review_service.set_publish_job_experiment(job_id, "")
    with pytest.raises(content_review_service.ContentReviewError, match="样本或官方导出周数还不足"):
        content_review_service.update_content_experiment(first_experiment_id, "keep")


def test_experiment_requires_twenty_treatment_twenty_baseline_and_three_export_weeks():
    account_id = _insert_account()
    _seed_diagnosis_baseline(account_id, captured_at="2026-08-29T12:00:00+08:00")
    recommendation = next(
        item
        for item in content_review_service.get_content_review_insights(account_id)["recommendations"]
        if item["diagnosis_code"] == "weak_opening"
    )
    created = content_review_service.create_content_experiment(
        account_id,
        recommendation["recommendation_id"],
    )
    experiment_id = created["experiment_id"]
    with get_connection() as connection:
        baseline = json.loads(
            connection.execute(
                "SELECT baseline_json FROM content_improvement_experiments WHERE id = ?",
                (experiment_id,),
            ).fetchone()[0]
        )
        baseline["work_count"] = 20
        connection.execute(
            "UPDATE content_improvement_experiments SET baseline_json = ? WHERE id = ?",
            (json.dumps(baseline, ensure_ascii=False), experiment_id),
        )
        connection.commit()

    treatment_items = []
    for index in range(20):
        published_at = f"2026-08-30T09:{index:02d}:00+08:00"
        title = f"实验门槛作品 {index}"
        job_id = _insert_publish_job(
            account_id,
            title=title,
            published_at=published_at,
            status="WAITING",
        )
        content_review_service.set_publish_job_experiment(job_id, experiment_id)
        treatment_items.append(
            _diagnosis_item(index + 200, title=title, published_at=published_at)
        )

    for week_index, captured_at in enumerate(
        [
            "2026-09-07T12:00:00+08:00",
            "2026-09-14T12:00:00+08:00",
            "2026-09-21T12:00:00+08:00",
        ]
    ):
        weekly_items = [
            {**item, "play_count": int(item["play_count"]) + week_index}
            for item in treatment_items
        ]
        content_review_service.commit_douyin_item_export(
            account_id=account_id,
            items=weekly_items,
            captured_at=captured_at,
            source_filename="作品列表导出.xlsx",
        )

    experiment = next(
        item
        for item in content_review_service.list_content_experiments(account_id)
        if item["id"] == experiment_id
    )
    assert experiment["progress"]["treatment_count"] == 20
    assert experiment["progress"]["baseline_count"] == 20
    assert experiment["progress"]["official_export_weeks"] == 3
    assert experiment["progress"]["trend_visible"] is True
    assert experiment["progress"]["decision_ready"] is True
    decided = content_review_service.update_content_experiment(experiment_id, "inconclusive")
    assert decided == {
        "status": "completed",
        "decision": "inconclusive",
        "message": "实验结论已记录。",
    }
